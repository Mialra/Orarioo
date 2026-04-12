import logging
import random
import re
from datetime import datetime, timedelta
from io import BytesIO

from django.db import transaction
from django.db.models import Max, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response

from auditableEntity.audit import create_audit_entry, suppress_audit_events
from auditableEntity.models import AuditActionType
from classroom.models import Classroom
from common.drf import TeamScopedAuditableModelViewSet
from common.errors.exceptions import ValidationAppError
from common.export_utils import build_csv_response, sanitize_filename_stem
from group.models import Group
from schedule.algorithm import BasicScheduleGenerator, ScheduleGenerationError
from schedule.algorithm.constraints.hard import (
    group_daily_limit,
    session_preference_state,
    teacher_preference_state,
)
from schedule.algorithm.evaluator import ScheduleEvaluator
from schedule.algorithm.generator import ScheduleReplanner
from schedule.algorithm.slots import (
    STAGE_SLOT_WINDOWS,
    session_stage_code,
    slot_preference_key_from_datetime,
)
from schedule.constants import AUTO_GENERATED_OBSERVATION, SAVED_TIMETABLE_PREFIX
from schedule.models import Schedule
from schedule.serializers import ScheduleSerializer
from subject.models import SubjectTimePreferenceState
from teacher.models import Teacher, TeacherTimePreferenceState
from user.models import User

logger = logging.getLogger(__name__)

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class ScheduleViewSet(TeamScopedAuditableModelViewSet):
    """CRUD API for schedules."""

    DEFAULT_GENERATION_OPTIONS = {
        "recess_supervisors_preschool": 0,
        "recess_supervisors_primary": 0,
        "include_tc": True,
        "tc_capacity": 1,
    }

    queryset = Schedule.objects.all().select_related(
        "teacher", "classroom", "group", "subject"
    )
    serializer_class = ScheduleSerializer
    EXPORT_ENTITY_ORDER = ["teacher", "classroom", "group"]
    EXPORT_ENTITY_CONFIG = {
        "teacher": {
            "label": "Profesor",
            "model": Teacher,
            "field": "teacher",
        },
        "group": {
            "label": "Curso",
            "model": Group,
            "field": "group",
        },
        "classroom": {
            "label": "Aula",
            "model": Classroom,
            "field": "classroom",
        },
    }
    STAGE_TC_BREAK_SLOTS = {
        "preschool": [("10:30", "11:00"), ("13:30", "14:00")],
        "primary": [("11:30", "12:00")],
        "secondary": [("11:00", "11:30")],
    }
    DAY_NAME_TO_WEEKDAY = {
        "Lunes": 0,
        "Martes": 1,
        "Miércoles": 2,
        "Jueves": 3,
        "Viernes": 4,
    }
    WEEKDAY_TO_DAY_NAME = {value: key for key, value in DAY_NAME_TO_WEEKDAY.items()}

    @staticmethod
    def _build_teacher_workloads(schedules):
        workloads = {}

        for schedule in schedules or []:
            teacher = getattr(schedule, "teacher", None)
            start_time = getattr(schedule, "start_time", None)
            end_time = getattr(schedule, "end_time", None)

            if teacher is None or start_time is None or end_time is None:
                continue

            duration_seconds = (end_time - start_time).total_seconds()
            if duration_seconds <= 0:
                continue

            duration_minutes = int(round(duration_seconds / 60.0))
            if duration_minutes <= 0:
                continue

            teacher_name = (getattr(teacher, "name", "") or "").strip()
            if not teacher_name:
                teacher_name = f"Profesor {teacher.id}"

            item = workloads.setdefault(
                teacher.id,
                {
                    "teacher_id": teacher.id,
                    "teacher_name": teacher_name,
                    "total_minutes": 0,
                },
            )
            item["total_minutes"] += duration_minutes

        return [
            {
                **item,
                "total_hours": round(item["total_minutes"] / 60.0, 2),
            }
            for item in sorted(
                workloads.values(),
                key=lambda value: value["teacher_name"].lower(),
            )
        ]

    @staticmethod
    def _parse_positive_int(raw_value, field_name):
        if raw_value in (None, ""):
            return None, None
        try:
            parsed = int(raw_value)
        except (TypeError, ValueError):
            return None, Response(
                {"detail": f"{field_name} must be a positive integer."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if parsed <= 0:
            return None, Response(
                {"detail": f"{field_name} must be a positive integer."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return parsed, None

    @staticmethod
    def _resolve_source_queryset(queryset, source):
        if source == "generated":
            return queryset.filter(observations=AUTO_GENERATED_OBSERVATION)
        if source == "saved":
            return queryset.exclude(observations=AUTO_GENERATED_OBSERVATION)
        return queryset

    @staticmethod
    def _resolve_entity_filtered_queryset(queryset, entity_type, entity_id):
        if entity_type == "group":
            return queryset.filter(group_id=entity_id)
        if entity_type == "teacher":
            return queryset.filter(teacher_id=entity_id)
        if entity_type == "classroom":
            return queryset.filter(classroom_id=entity_id)
        if entity_type == "subject":
            return queryset.filter(subject_id=entity_id)
        return queryset

    @staticmethod
    def _parse_bool_param(raw_value, field_name):
        if raw_value in (None, ""):
            return False, None

        normalized = str(raw_value).strip().lower()
        if normalized in {"1", "true", "yes", "on"}:
            return True, None
        if normalized in {"0", "false", "no", "off"}:
            return False, None

        return False, Response(
            {
                "detail": (
                    f"{field_name} must be a boolean value "
                    "(true/false, 1/0, yes/no)."
                )
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    @classmethod
    def _parse_id_list_param(cls, request, field_name):
        raw_values = []
        raw_single = request.query_params.get(field_name)
        if raw_single not in (None, ""):
            raw_values.extend(str(raw_single).split(","))

        for raw_value in request.query_params.getlist(field_name):
            if raw_value in (None, ""):
                continue
            raw_values.extend(str(raw_value).split(","))

        normalized = []
        for raw in raw_values:
            token = str(raw).strip()
            if not token:
                continue
            parsed, parse_error = cls._parse_positive_int(token, field_name)
            if parse_error is not None:
                return None, parse_error
            normalized.append(parsed)

        return sorted(set(normalized)), None

    @classmethod
    def _parse_card_filters(cls, request):
        card_specs = {
            "group": {"all_param": "group_all", "ids_param": "group_ids"},
            "teacher": {"all_param": "teacher_all", "ids_param": "teacher_ids"},
            "classroom": {
                "all_param": "classroom_all",
                "ids_param": "classroom_ids",
            },
        }

        filters = {}
        for entity_type, spec in card_specs.items():
            include_all, include_all_error = cls._parse_bool_param(
                request.query_params.get(spec["all_param"]),
                spec["all_param"],
            )
            if include_all_error is not None:
                return None, include_all_error

            selected_ids, selected_ids_error = cls._parse_id_list_param(
                request,
                spec["ids_param"],
            )
            if selected_ids_error is not None:
                return None, selected_ids_error

            filters[entity_type] = {
                "include_all": include_all,
                "ids": selected_ids,
            }

        has_any_filter = any(
            value["include_all"] or value["ids"] for value in filters.values()
        )
        return {
            "mode": "cards",
            "filters": filters,
            "has_any_filter": has_any_filter,
        }, None

    @staticmethod
    def _filter_queryset_with_cards(queryset, filters):
        card_to_field = {
            "group": "group",
            "teacher": "teacher",
            "classroom": "classroom",
        }

        criteria = Q(pk__in=[])
        has_criteria = False
        for entity_type, config in filters.items():
            field_name = card_to_field[entity_type]

            if config["include_all"]:
                criteria |= Q(**{f"{field_name}__isnull": False})
                has_criteria = True

            if config["ids"]:
                criteria |= Q(**{f"{field_name}_id__in": config["ids"]})
                has_criteria = True

        if not has_criteria:
            return queryset.none()

        return queryset.filter(criteria).distinct()

    @classmethod
    def _parse_export_params(cls, request):
        export_format = (
            (request.query_params.get("export_format") or "csv").strip().lower()
        )
        if export_format not in {"csv", "pdf"}:
            return None, Response(
                {"detail": "export_format must be one of: csv, pdf."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        source = (request.query_params.get("source") or "all").strip().lower()
        if source not in {"all", "generated", "saved"}:
            return None, Response(
                {"detail": "source must be one of: all, generated, saved."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        saved_timetable_name = (
            request.query_params.get("saved_timetable_name") or ""
        ).strip()

        if request.query_params.get("selection_mode") == "cards":
            card_filters, card_filter_error = cls._parse_card_filters(request)
            if card_filter_error is not None:
                return None, card_filter_error

            return {
                "format": export_format,
                "source": source,
                "scope": "cards",
                "card_filters": card_filters,
                "saved_timetable_name": saved_timetable_name,
            }, None

        scope = (request.query_params.get("scope") or "all").strip().lower()
        if scope not in {"all", "entity"}:
            return None, Response(
                {"detail": "scope must be one of: all, entity."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        entity_type = (request.query_params.get("entity_type") or "").strip().lower()
        entity_id = request.query_params.get("entity_id")

        if scope == "entity":
            if entity_type not in {"group", "teacher", "classroom", "subject"}:
                return None, Response(
                    {
                        "detail": (
                            "entity_type must be one of: group, teacher, classroom, "
                            "subject when scope=entity."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if entity_id in (None, ""):
                return None, Response(
                    {"detail": "entity_id is required when scope=entity."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            parsed_entity_id, entity_error = cls._parse_positive_int(
                entity_id,
                "entity_id",
            )
            if entity_error is not None:
                return None, entity_error
            entity_id = parsed_entity_id
        else:
            entity_type = None
            entity_id = None

        return {
            "format": export_format,
            "source": source,
            "scope": scope,
            "entity_type": entity_type,
            "entity_id": entity_id,
            "saved_timetable_name": saved_timetable_name,
        }, None

    @classmethod
    def _resolve_saved_schedule_name(cls, params, queryset):
        explicit_name = (params.get("saved_timetable_name") or "").strip()
        if explicit_name:
            return explicit_name

        if params.get("source") != "saved":
            return ""

        names = list(
            queryset.exclude(name__isnull=True)
            .exclude(name__exact="")
            .values_list("name", flat=True)
            .distinct()[:1]
        )
        if names:
            return names[0]
        return ""

    @classmethod
    def _build_export_filename(cls, params, saved_schedule_name=""):
        if params["source"] == "saved" and saved_schedule_name:
            stem = sanitize_filename_stem(saved_schedule_name, "orarioo_saved_schedule")
        else:
            date_token = timezone.now().strftime("%Y%m%d_%H%M%S")
            stem = f"orarioo_generated_schedule_{date_token}"
        return f"{stem}.{params['format']}"

    @classmethod
    def _build_export_rows(cls, schedules):
        weekday_to_name = {
            0: "Lunes",
            1: "Martes",
            2: "Miércoles",
            3: "Jueves",
            4: "Viernes",
            5: "Sábado",
            6: "Domingo",
        }

        rows = []
        for schedule in schedules:
            local_start = (
                timezone.localtime(schedule.start_time) if schedule.start_time else None
            )
            local_end = (
                timezone.localtime(schedule.end_time) if schedule.end_time else None
            )
            rows.append(
                {
                    "day": (
                        weekday_to_name.get(local_start.weekday(), "")
                        if local_start
                        else ""
                    ),
                    "start": local_start.strftime("%H:%M") if local_start else "",
                    "end": local_end.strftime("%H:%M") if local_end else "",
                    "subject": schedule.subject.name if schedule.subject else "",
                    "teacher": schedule.teacher.name if schedule.teacher else "",
                    "group": schedule.group.name if schedule.group else "",
                    "classroom": schedule.classroom.name if schedule.classroom else "",
                }
            )
        return rows

    @classmethod
    def _build_export_units(cls, queryset, params, active_team):
        units = []
        if params.get("scope") != "cards":
            rows = cls._build_export_rows(queryset)
            units.append(
                {
                    "entity_type": "mixed",
                    "header": "",
                    "rows": rows,
                    "schedules": list(queryset),
                }
            )
            return units

        filters = params["card_filters"]["filters"]
        for entity_type in cls.EXPORT_ENTITY_ORDER:
            config = cls.EXPORT_ENTITY_CONFIG[entity_type]
            field_name = config["field"]
            entity_filter = filters[entity_type]

            selected_ids = []
            if entity_filter["include_all"]:
                selected_ids.extend(
                    list(
                        queryset.exclude(**{f"{field_name}_id__isnull": True})
                        .values_list(f"{field_name}_id", flat=True)
                        .distinct()
                    )
                )
            selected_ids.extend(entity_filter["ids"])

            selected_ids = sorted(set(selected_ids))
            if not selected_ids:
                continue

            model_cls = config["model"]
            name_map = {
                obj.id: obj.name
                for obj in model_cls.objects.filter(
                    id__in=selected_ids,
                    team=active_team,
                ).only("id", "name")
            }

            for object_id in selected_ids:
                object_name = name_map.get(object_id, f"{config['label']} {object_id}")
                object_queryset = queryset.filter(
                    **{f"{field_name}_id": object_id}
                ).order_by("start_time", "id")
                units.append(
                    {
                        "entity_type": entity_type,
                        "header": f"{config['label']} {object_name}",
                        "rows": cls._build_export_rows(object_queryset),
                        "schedules": list(object_queryset),
                    }
                )

        return units

    @staticmethod
    def _build_csv_response(rows, filename):
        header = [
            "Día",
            "Inicio",
            "Fin",
            "Asignatura",
            "Profesor",
            "Curso",
            "Aula",
        ]
        return build_csv_response(
            header,
            [
                [
                    row["day"],
                    row["start"],
                    row["end"],
                    row["subject"],
                    row["teacher"],
                    row["group"],
                    row["classroom"],
                ]
                for row in rows
            ],
            filename,
        )

    @staticmethod
    def _make_unique_sheet_title(base_title, used_titles):
        """Generate a unique Excel sheet title (max 31 chars)."""
        safe = re.sub(r"[\\/*?:\[\]]+", "_", base_title).strip() or "Horario"
        safe = safe[:31]
        if safe not in used_titles:
            used_titles.add(safe)
            return safe
        counter = 2
        while True:
            suffix = f"_{counter}"
            candidate = f"{safe[:31 - len(suffix)]}{suffix}"
            if candidate not in used_titles:
                used_titles.add(candidate)
                return candidate
            counter += 1

    @staticmethod
    def _populate_excel_sheet(sheet, headers, rows):
        """Add headers and rows to an Excel sheet."""
        sheet.append(headers)
        for row in rows:
            sheet.append(
                [
                    row["day"],
                    row["start"],
                    row["end"],
                    row["subject"],
                    row["teacher"],
                    row["group"],
                    row["classroom"],
                ]
            )

    @classmethod
    def _build_excel_workbook(cls, units):
        """Create Excel workbook with multiple sheets from export units."""
        from openpyxl import Workbook

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        headers = ["Día", "Inicio", "Fin", "Asignatura", "Profesor", "Curso", "Aula"]
        used_titles = set()

        if not units:
            sheet = workbook.create_sheet(title="Sin datos")
            sheet.append(headers)
        else:
            for index, unit in enumerate(units, start=1):
                base_title = unit["header"] or f"Horario {index}"
                sheet_title = cls._make_unique_sheet_title(base_title, used_titles)
                sheet = workbook.create_sheet(title=sheet_title)
                cls._populate_excel_sheet(sheet, headers, unit["rows"])

        return workbook

    @classmethod
    def _build_excel_response(cls, units, filename):
        """Build Excel HTTP response with multiple sheets or fallback to CSV."""
        try:
            workbook = cls._build_excel_workbook(units)
        except ImportError:
            merged_rows = []
            for unit in units:
                merged_rows.extend(unit["rows"])
            fallback_filename = re.sub(r"\.xlsx$", ".csv", filename)
            return cls._build_csv_response(merged_rows, fallback_filename)

        buffer = BytesIO()
        workbook.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument." "spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @staticmethod
    def _normalize_stage(stage_value):
        """Normalize stage name to standard form."""
        value = (stage_value or "").strip().lower()
        if "preschool" in value or "infantil" in value:
            return "preschool"
        if "primary" in value or "primaria" in value:
            return "primary"
        if "secondary" in value or "eso" in value:
            return "secondary"
        return value

    @staticmethod
    def _describe_schedule(schedule):
        """Get display name for a schedule's subject."""
        return schedule.subject.name if schedule.subject else "-"

    @classmethod
    def _collect_slots_and_content(cls, schedules):
        """Collect all time slots and cell content from schedules."""
        weekday_to_name = {
            0: "Lunes",
            1: "Martes",
            2: "Miércoles",
            3: "Jueves",
            4: "Viernes",
        }
        slot_keys = []
        cell_content = {}
        day_stage_map = {}

        for schedule in schedules:
            local_start = timezone.localtime(schedule.start_time)
            local_end = timezone.localtime(schedule.end_time)
            weekday = local_start.weekday()
            if weekday not in weekday_to_name:
                continue

            day_name = weekday_to_name[weekday]
            slot = (local_start.strftime("%H:%M"), local_end.strftime("%H:%M"))
            if slot not in slot_keys:
                slot_keys.append(slot)

            key = (day_name, slot)
            cell_content.setdefault(key, []).append(cls._describe_schedule(schedule))

            # Track stages for TC breaks
            normalized_stage = cls._normalize_stage(
                getattr(schedule.group, "stage", "") if schedule.group else ""
            )
            if normalized_stage in cls.STAGE_TC_BREAK_SLOTS:
                day_stage_map.setdefault(day_name, set()).add(normalized_stage)

        return slot_keys, cell_content, day_stage_map

    @classmethod
    def _inject_tc_breaks(cls, slot_keys, cell_content, day_stage_map):
        """Add TC break slots to timetable for teacher schedules."""
        for day_name, stages in day_stage_map.items():
            for stage in stages:
                for tc_slot in cls.STAGE_TC_BREAK_SLOTS[stage]:
                    if tc_slot not in slot_keys:
                        slot_keys.append(tc_slot)
                    key = (day_name, tc_slot)
                    cell_content.setdefault(key, []).append("Trabajo de Centro")

    @classmethod
    def _build_timetable_rows(cls, slot_keys, cell_content):
        """Build table rows from slots and content."""
        days = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
        table_data = [["Hora", *days]]
        for slot in slot_keys:
            row = [f"{slot[0]} - {slot[1]}"]
            for day in days:
                entries = cell_content.get((day, slot), [])
                row.append("\n\n".join(entries))
            table_data.append(row)
        return table_data

    @classmethod
    def _build_timetable_table_data(cls, schedules, entity_type):
        """Build table data for timetable PDF/export."""
        slot_keys, cell_content, day_stage_map = cls._collect_slots_and_content(
            schedules
        )

        if entity_type == "teacher":
            cls._inject_tc_breaks(slot_keys, cell_content, day_stage_map)

        slot_keys.sort()
        if not slot_keys:
            return [
                ["Hora", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes"],
                ["Sin sesiones", "", "", "", "", ""],
            ]

        return cls._build_timetable_rows(slot_keys, cell_content)

    @staticmethod
    def _build_pdf_response(rows, filename, title_text=""):
        if not REPORTLAB_AVAILABLE:
            return Response(
                {
                    "detail": (
                        "PDF export is unavailable because reportlab is not installed. "
                        "Install reportlab to enable PDF exports."
                    )
                },
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=24,
            rightMargin=24,
            topMargin=24,
            bottomMargin=24,
        )

        del rows
        del title_text
        story = []
        document.build(story)

        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @classmethod
    def _build_pdf_units_response(cls, units, filename):
        if not REPORTLAB_AVAILABLE:
            return Response(
                {
                    "detail": (
                        "PDF export is unavailable because reportlab is not installed. "
                        "Install reportlab to enable PDF exports."
                    )
                },
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=24,
            rightMargin=24,
            topMargin=24,
            bottomMargin=24,
        )

        styles = getSampleStyleSheet()
        story = []

        for index, unit in enumerate(units):
            if index > 0:
                story.append(PageBreak())

            if unit["header"]:
                story.extend(
                    [
                        Paragraph(f"<b>{unit['header']}</b>", styles["Title"]),
                        Spacer(1, 10),
                    ]
                )

            table_data = cls._build_timetable_table_data(
                unit["schedules"], unit["entity_type"]
            )

            available_width = document.width
            time_col_width = available_width * 0.15
            day_col_width = (available_width - time_col_width) / 5
            col_widths = [time_col_width, *([day_col_width] * 5)]

            table = Table(table_data, repeatRows=1, colWidths=col_widths)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e7ecfb")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#b7bfd4")),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 4),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            story.append(table)

        document.build(story)

        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def export(self, request):
        params, error_response = self._parse_export_params(request)
        if error_response is not None:
            return error_response

        queryset = self.get_queryset().order_by("start_time", "id")
        queryset = self._resolve_source_queryset(queryset, params["source"])
        if params["scope"] == "entity":
            queryset = self._resolve_entity_filtered_queryset(
                queryset,
                params["entity_type"],
                params["entity_id"],
            )
        elif params["scope"] == "cards":
            queryset = self._filter_queryset_with_cards(
                queryset,
                params["card_filters"]["filters"],
            )

        saved_schedule_name = self._resolve_saved_schedule_name(params, queryset)
        units = self._build_export_units(
            queryset,
            params,
            active_team=self.get_active_team(),
        )
        rows = self._build_export_rows(queryset)
        filename = self._build_export_filename(params, saved_schedule_name)

        if params["format"] == "csv":
            if params.get("scope") == "cards":
                excel_filename = filename.rsplit(".", 1)[0] + ".xlsx"
                return self._build_excel_response(units, excel_filename)
            return self._build_csv_response(rows, filename)
        return self._build_pdf_units_response(units, filename)

    @staticmethod
    def _parse_generation_int(payload, field_name, *, min_value, max_value):
        raw_value = payload.get(field_name)
        if raw_value in (None, ""):
            return None, None
        try:
            value = int(raw_value)
        except (TypeError, ValueError):
            return None, Response(
                {"detail": f"{field_name} must be an integer value."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if value < min_value or value > max_value:
            return None, Response(
                {
                    "detail": (
                        f"{field_name} must be between {min_value} and {max_value}."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        return value, None

    @classmethod
    def _parse_generation_bool(cls, payload, field_name):
        raw_value = payload.get(field_name)
        if raw_value in (None, ""):
            return None, None

        if isinstance(raw_value, bool):
            return raw_value, None

        return cls._parse_bool_param(raw_value, field_name)

    @classmethod
    def _parse_base_generation_int_options(cls, payload, options):
        int_fields = {
            "recess_supervisors_preschool": (0, 20),
            "recess_supervisors_primary": (0, 20),
        }

        if options.get("include_tc", True):
            int_fields["tc_capacity"] = (1, 10)

        for field_name, bounds in int_fields.items():
            parsed, error_response = cls._parse_generation_int(
                payload,
                field_name,
                min_value=bounds[0],
                max_value=bounds[1],
            )
            if error_response is not None:
                return error_response
            if parsed is not None:
                options[field_name] = parsed
        return None

    @classmethod
    def _parse_base_generation_bool_options(cls, payload, options):
        bool_fields = ["include_tc"]

        for field_name in bool_fields:
            parsed, error_response = cls._parse_generation_bool(payload, field_name)
            if error_response is not None:
                return error_response
            if parsed is not None:
                options[field_name] = parsed
        return None

    @classmethod
    def _parse_generation_options(cls, payload):
        options = dict(cls.DEFAULT_GENERATION_OPTIONS)
        bool_options_error = cls._parse_base_generation_bool_options(
            payload,
            options,
        )
        if bool_options_error is not None:
            return None, bool_options_error

        base_options_error = cls._parse_base_generation_int_options(
            payload,
            options,
        )
        if base_options_error is not None:
            return None, base_options_error

        return options, None

    def generate(self, request):
        actor = getattr(request.user, "email", "")
        active_team = self.get_active_team()
        raw_seed = request.data.get("seed")
        generation_options, options_error = self._parse_generation_options(request.data)
        if options_error is not None:
            raise ValidationAppError(
                "INVALID_GENERATION_OPTION",
                options_error.data.get("detail", "Invalid schedule generation option."),
            )

        if raw_seed in (None, ""):
            generation_seed = random.SystemRandom().randrange(1, 2**31 - 1)
        else:
            try:
                generation_seed = int(raw_seed)
            except (TypeError, ValueError):
                raise ValidationAppError(
                    "INVALID_INTEGER",
                    "seed must be an integer value.",
                    field_name="seed",
                    context={"field": "seed", "value": raw_seed},
                )

        try:
            schedules = BasicScheduleGenerator.generate(
                actor_email=actor,
                user=request.user,
                team=active_team,
                random_seed=generation_seed,
                generation_options=generation_options,
            )
        except ScheduleGenerationError as exc:
            logger.warning(
                "Schedule generation rejected: actor=%s, reason=%s",
                actor,
                exc,
            )
            raise
        serialized = self.get_serializer(schedules, many=True)
        teacher_workloads = self._build_teacher_workloads(schedules)

        return Response(
            {
                "detail": "Schedule generated successfully.",
                "seed": generation_seed,
                "generation_options": generation_options,
                "schedules": serialized.data,
                "generated_count": len(serialized.data),
                "teacher_workloads": teacher_workloads,
            },
            status=status.HTTP_201_CREATED,
        )

    def _saved_queryset_for_user(self, request_user):
        return (
            self.get_queryset()
            .exclude(observations=AUTO_GENERATED_OBSERVATION)
            .filter(users=request_user)
        )

    def saved(self, request):
        saved_queryset = self._saved_queryset_for_user(request.user).order_by(
            "start_time", "id"
        )
        saved_schedules = list(saved_queryset)
        serialized = self.get_serializer(saved_schedules, many=True)
        return Response(
            {
                "count": len(serialized.data),
                "results": serialized.data,
                "teacher_workloads": self._build_teacher_workloads(saved_schedules),
            },
            status=status.HTTP_200_OK,
        )

    def saved_summary(self, request):
        summary_queryset = (
            self._saved_queryset_for_user(request.user)
            .exclude(name__isnull=True)
            .exclude(name__exact="")
            .values("name")
            .annotate(updated_at=Max("updated_at"))
            .order_by("-updated_at", "name")
        )
        summary_items = list(summary_queryset)
        return Response(
            {
                "count": len(summary_items),
                "results": summary_items,
            },
            status=status.HTTP_200_OK,
        )

    def saved_detail(self, request):
        timetable_name = (request.query_params.get("timetable_name") or "").strip()
        if not timetable_name:
            return Response(
                {"detail": "timetable_name query parameter is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        schedules, error_response = self._fetch_saved_timetable_schedules(
            request_user=request.user,
            timetable_name=timetable_name,
            team=self.get_active_team(),
        )
        if error_response is not None:
            return error_response

        serialized = self.get_serializer(schedules, many=True)
        return Response(
            {
                "count": len(serialized.data),
                "results": serialized.data,
                "teacher_workloads": self._build_teacher_workloads(schedules),
            },
            status=status.HTTP_200_OK,
        )

    @staticmethod
    def _parse_saved_timetable_name(payload):
        timetable_name = (payload.get("timetable_name") or "").strip()
        if not timetable_name:
            return None, Response(
                {"timetable_name": "timetable_name is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return timetable_name, None

    @staticmethod
    def _fetch_saved_timetable_schedules(*, request_user, timetable_name, team):
        saved_observation = f"{SAVED_TIMETABLE_PREFIX}: {timetable_name}"
        schedules = list(
            Schedule.objects.filter(
                users=request_user,
                observations=saved_observation,
                team=team,
            ).order_by("id")
        )
        if not schedules:
            return None, Response(
                {"detail": "Saved timetable not found."},
                status=status.HTTP_404_NOT_FOUND,
            )
        return schedules, None

    @action(detail=False, methods=["post"], url_path="delete-saved-timetable")
    def delete_saved_timetable(self, request):
        timetable_name, error_response = self._parse_saved_timetable_name(request.data)
        if error_response is not None:
            return error_response

        schedules, error_response = self._fetch_saved_timetable_schedules(
            request_user=request.user,
            timetable_name=timetable_name,
            team=self.get_active_team(),
        )
        if error_response is not None:
            return error_response

        deleted_count = len(schedules)
        representative_schedule_id = schedules[0].pk
        with suppress_audit_events(("schedule", AuditActionType.DELETE)):
            for schedule in schedules:
                schedule.delete()

        create_audit_entry(
            model=Schedule,
            entity_id=representative_schedule_id,
            entity_name=timetable_name,
            action_type=AuditActionType.DELETE,
            detail=(
                f'Se elimino el horario guardado "{timetable_name}" '
                f"con {deleted_count} sesiones."
            ),
            changed_fields=[
                {
                    "campo": "Sesiones eliminadas",
                    "valor_anterior": deleted_count,
                }
            ],
            team=self.get_active_team(),
        )
        return Response(
            {
                "detail": "Saved timetable deleted successfully.",
                "deleted_count": deleted_count,
            },
            status=status.HTTP_200_OK,
        )

    @staticmethod
    def _parse_int_list(payload, field_name):
        raw_values = payload.get(field_name) or []
        if field_name == "schedule_ids":
            if not isinstance(raw_values, list) or not raw_values:
                return None, Response(
                    {"detail": "schedule_ids must be a non-empty list."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        elif not isinstance(raw_values, list):
            return None, Response(
                {"detail": f"{field_name} must be a list."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            normalized_values = [int(value) for value in raw_values]
        except (TypeError, ValueError):
            return None, Response(
                {"detail": f"{field_name} must contain integer values."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if any(value <= 0 for value in normalized_values):
            return None, Response(
                {field_name: f"{field_name} must contain positive integer values."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if len(set(normalized_values)) != len(normalized_values):
            return None, Response(
                {field_name: f"{field_name} cannot contain duplicated values."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return normalized_values, None

    @staticmethod
    def _ensure_request_user_in_user_ids(request_user_id, normalized_user_ids):
        if request_user_id not in normalized_user_ids:
            normalized_user_ids.append(request_user_id)
        return normalized_user_ids

    @staticmethod
    def _fetch_target_users(normalized_user_ids, active_team):
        requested_ids = set(normalized_user_ids)
        target_users = list(
            User.objects.filter(
                id__in=requested_ids,
                collaboration_teams=active_team,
            )
        )
        if len(target_users) != len(requested_ids):
            return None, Response(
                {
                    "detail": (
                        "Some user_ids do not exist or are outside the active team."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        return target_users, None

    @staticmethod
    def _fetch_eligible_schedules(normalized_ids, request_user, actor_email, team):
        requested_ids = set(normalized_ids)
        schedules = list(
            Schedule.objects.filter(
                id__in=normalized_ids,
                users=request_user,
                created_by=actor_email,
                observations=AUTO_GENERATED_OBSERVATION,
                team=team,
            )
        )
        if len(schedules) != len(requested_ids):
            return None, Response(
                {
                    "detail": (
                        "Some schedules were not found or are not eligible to be saved "
                        "(must belong to current user and be auto-generated)."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        return schedules, None

    @staticmethod
    def _persist_saved_schedules(
        *, schedules, timetable_name, actor_email, target_users
    ):
        saved_observation = f"{SAVED_TIMETABLE_PREFIX}: {timetable_name}"
        for schedule in schedules:
            schedule.name = timetable_name
            schedule.observations = saved_observation
            schedule.updated_by = actor_email
            schedule.save(
                update_fields=["name", "observations", "updated_by", "updated_at"]
            )
            schedule.users.add(*target_users)

    @staticmethod
    def _saved_timetable_name_exists(*, request_user, timetable_name, team):
        saved_observation = f"{SAVED_TIMETABLE_PREFIX}: {timetable_name}"
        return Schedule.objects.filter(
            users=request_user,
            observations=saved_observation,
            team=team,
        ).exists()

    def save_generated(self, request):
        actor = getattr(request.user, "email", "")
        active_team = self.get_active_team()
        timetable_name = (request.data.get("timetable_name") or "").strip()

        if not timetable_name:
            return Response(
                {"timetable_name": "timetable_name is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if self._saved_timetable_name_exists(
            request_user=request.user,
            timetable_name=timetable_name,
            team=active_team,
        ):
            return Response(
                {
                    "timetable_name": (
                        "A saved timetable with this name already exists. "
                        "Use another name or delete the previous one."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        normalized_ids, error_response = self._parse_int_list(
            request.data,
            "schedule_ids",
        )
        if error_response is not None:
            return error_response

        normalized_user_ids, error_response = self._parse_int_list(
            request.data,
            "user_ids",
        )
        if error_response is not None:
            return error_response

        normalized_user_ids = self._ensure_request_user_in_user_ids(
            request.user.id,
            normalized_user_ids,
        )

        target_users, error_response = self._fetch_target_users(
            normalized_user_ids,
            active_team,
        )
        if error_response is not None:
            return error_response

        schedules, error_response = self._fetch_eligible_schedules(
            normalized_ids,
            request.user,
            actor,
            active_team,
        )
        if error_response is not None:
            return error_response

        self._persist_saved_schedules(
            schedules=schedules,
            timetable_name=timetable_name,
            actor_email=actor,
            target_users=target_users,
        )

        serialized = self.get_serializer(schedules, many=True)
        return Response(
            {
                "detail": "Generated schedules saved successfully.",
                "saved_count": len(schedules),
                "schedules": serialized.data,
                "teacher_workloads": self._build_teacher_workloads(schedules),
            },
            status=status.HTTP_200_OK,
        )

    @staticmethod
    def _parse_hhmm(raw_value, field_name):
        value = (raw_value or "").strip()
        try:
            return datetime.strptime(value, "%H:%M").time(), None
        except ValueError:
            return None, Response(
                {"detail": f"{field_name} must follow HH:MM format."},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @classmethod
    def _normalize_move_mode(cls, raw_mode):
        mode = (raw_mode or "move").strip().lower()
        if mode not in {"move", "swap"}:
            return None, Response(
                {"detail": "mode must be one of: move, swap."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return mode, None

    @classmethod
    def _parse_move_slot(cls, slot_data, slot_label, *, require_schedule_id=False):
        if not isinstance(slot_data, dict):
            return None, Response(
                {"detail": f"{slot_label} must be an object."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        day_name = (slot_data.get("day") or "").strip()
        if day_name not in cls.DAY_NAME_TO_WEEKDAY:
            return None, Response(
                {
                    "detail": (
                        f"{slot_label}.day must be one of: "
                        "Lunes, Martes, Miércoles, Jueves, Viernes."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        start_raw = slot_data.get("start")
        end_raw = slot_data.get("end")
        start_time, start_error = cls._parse_hhmm(start_raw, f"{slot_label}.start")
        if start_error is not None:
            return None, start_error
        end_time, end_error = cls._parse_hhmm(end_raw, f"{slot_label}.end")
        if end_error is not None:
            return None, end_error
        if end_time <= start_time:
            return None, Response(
                {
                    "detail": f"{slot_label}.end must be greater than {slot_label}.start."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        schedule_id = None
        raw_schedule_id = slot_data.get("schedule_id")
        if require_schedule_id and raw_schedule_id in (None, ""):
            return None, Response(
                {"detail": f"{slot_label}.schedule_id is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if raw_schedule_id not in (None, ""):
            parsed_schedule_id, parse_error = cls._parse_positive_int(
                raw_schedule_id,
                f"{slot_label}.schedule_id",
            )
            if parse_error is not None:
                return None, parse_error
            schedule_id = parsed_schedule_id

        return {
            "day": day_name,
            "start": start_time.strftime("%H:%M"),
            "end": end_time.strftime("%H:%M"),
            "start_time": start_time,
            "end_time": end_time,
            "schedule_id": schedule_id,
        }, None

    @classmethod
    def _slot_descriptor_from_datetimes(cls, start_dt, end_dt):
        local_start = timezone.localtime(start_dt)
        local_end = timezone.localtime(end_dt)
        return {
            "day": cls.WEEKDAY_TO_DAY_NAME.get(local_start.weekday(), ""),
            "start": local_start.strftime("%H:%M"),
            "end": local_end.strftime("%H:%M"),
        }

    @classmethod
    def _resolve_slot_datetimes_for_source_week(
        cls,
        *,
        source_start,
        day_name,
        start_time,
        end_time,
    ):
        source_local = timezone.localtime(source_start)
        monday_date = source_local.date() - timedelta(days=source_local.weekday())
        target_date = monday_date + timedelta(days=cls.DAY_NAME_TO_WEEKDAY[day_name])
        current_tz = timezone.get_current_timezone()
        target_start = timezone.make_aware(
            datetime.combine(target_date, start_time),
            current_tz,
        )
        target_end = timezone.make_aware(
            datetime.combine(target_date, end_time),
            current_tz,
        )
        return target_start, target_end

    def _resolve_timetable_scope_queryset(
        self,
        *,
        request_user,
        source_schedule,
        active_team,
    ):
        scoped_queryset = self.get_queryset().filter(
            users=request_user,
            team=active_team,
        )
        if source_schedule.observations == AUTO_GENERATED_OBSERVATION:
            return scoped_queryset.filter(
                observations=AUTO_GENERATED_OBSERVATION,
                created_by=source_schedule.created_by,
            )
        if source_schedule.observations.startswith(f"{SAVED_TIMETABLE_PREFIX}:"):
            return scoped_queryset.filter(observations=source_schedule.observations)
        return scoped_queryset.filter(
            name=source_schedule.name,
            observations=source_schedule.observations,
        )

    @staticmethod
    def _times_overlap(*, left_start, left_end, right_start, right_end):
        return left_start < right_end and right_start < left_end

    @staticmethod
    def _normalize_clock(value):
        return value.replace(second=0, microsecond=0, tzinfo=None)

    def _is_stage_window_allowed(self, *, schedule, start_dt, end_dt):
        stage_code = session_stage_code(
            session={"group": schedule.group, "subject": schedule.subject}
        )
        allowed_windows = STAGE_SLOT_WINDOWS.get(stage_code, [])

        local_start = timezone.localtime(start_dt)
        local_end = timezone.localtime(end_dt)
        if local_start.date() != local_end.date() or local_start.weekday() > 4:
            return False

        candidate_window = (
            self._normalize_clock(local_start.time()),
            self._normalize_clock(local_end.time()),
        )
        normalized_allowed = {
            (self._normalize_clock(left), self._normalize_clock(right))
            for left, right in allowed_windows
        }
        return candidate_window in normalized_allowed

    @staticmethod
    def _validate_target_preferences(*, schedule, start_dt):
        slot_key = slot_preference_key_from_datetime(slot=start_dt)
        if slot_key is None:
            return None

        session_ctx = {"subject": schedule.subject, "teacher": schedule.teacher}
        subject_state = session_preference_state(
            session=session_ctx,
            slot_preference_key=slot_key,
        )
        if subject_state == SubjectTimePreferenceState.UNAVAILABLE:
            return f"Subject '{schedule.subject.name}' is unavailable at {slot_key}."

        teacher_state = teacher_preference_state(
            session=session_ctx,
            slot_preference_key=slot_key,
        )
        if teacher_state == TeacherTimePreferenceState.UNAVAILABLE:
            return f"Teacher '{schedule.teacher.name}' is unavailable at {slot_key}."
        return None

    @staticmethod
    def _build_hypothetical_times(*, scope_schedules, assignments):
        hypothetical = {
            schedule.id: (schedule.start_time, schedule.end_time)
            for schedule in scope_schedules
        }
        hypothetical.update(assignments)
        return hypothetical

    def _validate_resource_overlaps_for_changes(
        self,
        *,
        scope_schedules,
        hypothetical_times,
        changed_ids,
    ):
        schedule_by_id = {schedule.id: schedule for schedule in scope_schedules}
        for changed_id in changed_ids:
            current_schedule = schedule_by_id[changed_id]
            current_start, current_end = hypothetical_times[changed_id]
            for other_schedule in scope_schedules:
                if other_schedule.id == changed_id:
                    continue

                other_start, other_end = hypothetical_times[other_schedule.id]
                if not self._times_overlap(
                    left_start=current_start,
                    left_end=current_end,
                    right_start=other_start,
                    right_end=other_end,
                ):
                    continue

                if (
                    current_schedule.teacher_id is not None
                    and current_schedule.teacher_id == other_schedule.teacher_id
                ):
                    return Response(
                        {
                            "detail": (
                                "Teacher conflict detected in target slot for "
                                f"'{current_schedule.teacher.name}'."
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                if (
                    current_schedule.group_id is not None
                    and current_schedule.group_id == other_schedule.group_id
                ):
                    return Response(
                        {
                            "detail": (
                                "Group conflict detected in target slot for "
                                f"'{current_schedule.group.name}'."
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                if (
                    current_schedule.classroom_id is not None
                    and current_schedule.classroom_id == other_schedule.classroom_id
                ):
                    return Response(
                        {
                            "detail": (
                                "Classroom conflict detected in target slot for "
                                f"'{current_schedule.classroom.name}'."
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

        return None

    def _validate_group_daily_limits(
        self,
        *,
        scope_schedules,
        hypothetical_times,
        changed_group_ids,
    ):
        group_by_id = {}
        day_count_by_group = {}

        for schedule in scope_schedules:
            if schedule.group_id not in changed_group_ids:
                continue
            group_by_id[schedule.group_id] = schedule.group
            schedule_start, _ = hypothetical_times[schedule.id]
            schedule_day = timezone.localtime(schedule_start).date()
            key = (schedule.group_id, schedule_day)
            day_count_by_group[key] = day_count_by_group.get(key, 0) + 1

        for (group_id, _), count in day_count_by_group.items():
            group = group_by_id.get(group_id)
            if group is None:
                continue
            if count > group_daily_limit(group):
                return Response(
                    {
                        "detail": (
                            f"Group '{group.name}' exceeds daily slot limit for "
                            "its stage."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
        return None

    def _validate_group_intraday_gaps(
        self,
        *,
        scope_schedules,
        hypothetical_times,
        changed_group_ids,
    ):
        schedules_by_group = self._group_schedules_by_id(
            scope_schedules=scope_schedules,
            changed_group_ids=changed_group_ids,
        )

        for group_schedules in schedules_by_group.values():
            if not group_schedules:
                continue

            reference_group = group_schedules[0].group
            window_to_index = self._window_index_by_stage(reference_group)
            if not window_to_index:
                continue

            by_day_indices = self._collect_group_day_window_indices(
                group_schedules=group_schedules,
                hypothetical_times=hypothetical_times,
                window_to_index=window_to_index,
            )
            if by_day_indices is None:
                continue

            for occupied_indices in by_day_indices.values():
                if self._has_intraday_gap(occupied_indices):
                    return Response(
                        {
                            "detail": (
                                f"Group '{reference_group.name}' would have intraday "
                                "gaps with that move."
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

        return None

    @staticmethod
    def _group_schedules_by_id(*, scope_schedules, changed_group_ids):
        grouped = {}
        for schedule in scope_schedules:
            if schedule.group_id in changed_group_ids:
                grouped.setdefault(schedule.group_id, []).append(schedule)
        return grouped

    def _window_index_by_stage(self, group):
        stage_code = session_stage_code(session={"group": group, "subject": None})
        allowed_windows = STAGE_SLOT_WINDOWS.get(stage_code, [])
        return {
            (self._normalize_clock(left), self._normalize_clock(right)): index
            for index, (left, right) in enumerate(allowed_windows)
        }

    def _collect_group_day_window_indices(
        self,
        *,
        group_schedules,
        hypothetical_times,
        window_to_index,
    ):
        by_day = {}
        for schedule in group_schedules:
            start_dt, end_dt = hypothetical_times[schedule.id]
            day_key = timezone.localtime(start_dt).date()
            by_day.setdefault(day_key, []).append((start_dt, end_dt))

        by_day_indices = {}
        for day_key, day_items in by_day.items():
            occupied_indices = []
            for start_dt, end_dt in day_items:
                local_start = timezone.localtime(start_dt)
                local_end = timezone.localtime(end_dt)
                window_key = (
                    self._normalize_clock(local_start.time()),
                    self._normalize_clock(local_end.time()),
                )
                index = window_to_index.get(window_key)
                if index is None:
                    return None
                occupied_indices.append(index)
            if occupied_indices:
                by_day_indices[day_key] = occupied_indices
        return by_day_indices

    @staticmethod
    def _has_intraday_gap(occupied_indices):
        first_idx = min(occupied_indices)
        last_idx = max(occupied_indices)
        occupied_set = set(occupied_indices)
        return any(
            index not in occupied_set for index in range(first_idx, last_idx + 1)
        )

    def _validate_minimal_move_constraints(
        self,
        *,
        scope_schedules,
        assignments,
        changed_ids,
    ):
        hypothetical_times = self._build_hypothetical_times(
            scope_schedules=scope_schedules,
            assignments=assignments,
        )
        schedule_by_id = {schedule.id: schedule for schedule in scope_schedules}

        for changed_id in changed_ids:
            schedule = schedule_by_id[changed_id]
            start_dt, end_dt = hypothetical_times[changed_id]

            if end_dt <= start_dt:
                return Response(
                    {"detail": "Target slot must end after it starts."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if not self._is_stage_window_allowed(
                schedule=schedule,
                start_dt=start_dt,
                end_dt=end_dt,
            ):
                return Response(
                    {
                        "detail": (
                            "Target slot is not allowed for the session stage "
                            f"({schedule.group.stage})."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            preference_error = self._validate_target_preferences(
                schedule=schedule,
                start_dt=start_dt,
            )
            if preference_error is not None:
                return Response(
                    {"detail": preference_error},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        overlap_error = self._validate_resource_overlaps_for_changes(
            scope_schedules=scope_schedules,
            hypothetical_times=hypothetical_times,
            changed_ids=changed_ids,
        )
        if overlap_error is not None:
            return overlap_error

        changed_group_ids = {
            schedule_by_id[schedule_id].group_id
            for schedule_id in changed_ids
            if schedule_by_id[schedule_id].group_id is not None
        }

        daily_limit_error = self._validate_group_daily_limits(
            scope_schedules=scope_schedules,
            hypothetical_times=hypothetical_times,
            changed_group_ids=changed_group_ids,
        )
        if daily_limit_error is not None:
            return daily_limit_error

        gap_error = self._validate_group_intraday_gaps(
            scope_schedules=scope_schedules,
            hypothetical_times=hypothetical_times,
            changed_group_ids=changed_group_ids,
        )
        if gap_error is not None:
            return gap_error

        return None

    def _fetch_source_schedule_for_move(
        self, *, request_user, active_team, source_slot
    ):
        source_schedule = (
            self.get_queryset()
            .filter(
                id=source_slot["schedule_id"],
                users=request_user,
                team=active_team,
            )
            .first()
        )
        if source_schedule is None:
            return None, Response(
                {"detail": "Source schedule not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        source_local_start = timezone.localtime(source_schedule.start_time)
        source_local_end = timezone.localtime(source_schedule.end_time)
        actual_source_day = self.WEEKDAY_TO_DAY_NAME.get(source_local_start.weekday())
        source_outdated = (
            actual_source_day != source_slot["day"]
            or source_local_start.strftime("%H:%M") != source_slot["start"]
            or source_local_end.strftime("%H:%M") != source_slot["end"]
        )
        if source_outdated:
            return None, Response(
                {
                    "detail": (
                        "The source slot no longer matches current data. "
                        "Refresh and try again."
                    )
                },
                status=status.HTTP_409_CONFLICT,
            )

        return source_schedule, None

    @staticmethod
    def _move_no_changes_response(mode):
        return Response(
            {
                "detail": "No changes were applied.",
                "mode": mode,
                "no_changes": True,
                "affected_schedules": [],
                "affected_slots": [],
                "teacher_workloads": [],
            },
            status=status.HTTP_200_OK,
        )

    def _resolve_move_scope(self, *, request_user, source_schedule, active_team):
        scope_queryset = self._resolve_timetable_scope_queryset(
            request_user=request_user,
            source_schedule=source_schedule,
            active_team=active_team,
        )
        scope_schedules = list(scope_queryset.order_by("id"))
        scope_by_id = {schedule.id: schedule for schedule in scope_schedules}
        if source_schedule.id not in scope_by_id:
            return (
                None,
                None,
                None,
                Response(
                    {"detail": "Source schedule is outside editable timetable scope."},
                    status=status.HTTP_400_BAD_REQUEST,
                ),
            )
        return scope_queryset, scope_schedules, scope_by_id, None

    def _resolve_swap_target_for_move(
        self,
        *,
        scope_queryset,
        scope_by_id,
        source_schedule,
        target_slot,
        target_start_dt,
        target_end_dt,
    ):
        target_schedule_id = target_slot["schedule_id"]
        if target_schedule_id is not None:
            target_schedule = scope_by_id.get(target_schedule_id)
            if target_schedule is None:
                return None, Response(
                    {
                        "detail": (
                            "target_slot.schedule_id must belong to the same "
                            "timetable scope as source_slot.schedule_id."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if (
                target_schedule.start_time != target_start_dt
                or target_schedule.end_time != target_end_dt
            ):
                return None, Response(
                    {
                        "detail": (
                            "Target schedule no longer matches target slot. "
                            "Refresh and try again."
                        )
                    },
                    status=status.HTTP_409_CONFLICT,
                )
            return target_schedule, None

        target_schedule = (
            scope_queryset.filter(start_time=target_start_dt, end_time=target_end_dt)
            .exclude(id=source_schedule.id)
            .order_by("id")
            .first()
        )
        if target_schedule is None:
            return None, Response(
                {"detail": "Swap requires a target schedule in destination slot."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return target_schedule, None

    @staticmethod
    def _apply_move_assignments(
        *,
        source_schedule,
        target_schedule,
        target_start_dt,
        target_end_dt,
        original_source_times,
        actor,
    ):
        affected_schedules = []
        with transaction.atomic():
            source_schedule.start_time = target_start_dt
            source_schedule.end_time = target_end_dt
            source_schedule.updated_by = actor
            source_schedule.save(
                update_fields=["start_time", "end_time", "updated_by", "updated_at"]
            )
            affected_schedules.append(source_schedule)

            if target_schedule is not None:
                target_schedule.start_time = original_source_times[0]
                target_schedule.end_time = original_source_times[1]
                target_schedule.updated_by = actor
                target_schedule.save(
                    update_fields=[
                        "start_time",
                        "end_time",
                        "updated_by",
                        "updated_at",
                    ]
                )
                affected_schedules.append(target_schedule)
        return affected_schedules

    def _build_affected_slot_descriptors(
        self,
        *,
        original_source_times,
        target_start_dt,
        target_end_dt,
        original_target_times,
    ):
        affected_slots = [
            self._slot_descriptor_from_datetimes(
                original_source_times[0],
                original_source_times[1],
            ),
            self._slot_descriptor_from_datetimes(target_start_dt, target_end_dt),
        ]
        if original_target_times is not None:
            affected_slots.append(
                self._slot_descriptor_from_datetimes(
                    original_target_times[0],
                    original_target_times[1],
                )
            )

        unique_affected_slots = []
        seen_slots = set()
        for slot in affected_slots:
            key = (slot["day"], slot["start"], slot["end"])
            if key in seen_slots:
                continue
            seen_slots.add(key)
            unique_affected_slots.append(slot)
        return unique_affected_slots

    def _parse_move_request_payload(self, payload):
        mode, mode_error = self._normalize_move_mode(payload.get("mode"))
        if mode_error is not None:
            return None, mode_error

        source_slot, source_error = self._parse_move_slot(
            payload.get("source_slot"),
            "source_slot",
            require_schedule_id=True,
        )
        if source_error is not None:
            return None, source_error

        target_slot, target_error = self._parse_move_slot(
            payload.get("target_slot"),
            "target_slot",
            require_schedule_id=False,
        )
        if target_error is not None:
            return None, target_error

        return {
            "mode": mode,
            "source_slot": source_slot,
            "target_slot": target_slot,
        }, None

    def _resolve_target_schedule_for_mode(
        self,
        *,
        mode,
        scope_queryset,
        scope_by_id,
        source_schedule,
        target_slot,
        target_start_dt,
        target_end_dt,
    ):
        if mode != "swap":
            return None, None

        return self._resolve_swap_target_for_move(
            scope_queryset=scope_queryset,
            scope_by_id=scope_by_id,
            source_schedule=source_schedule,
            target_slot=target_slot,
            target_start_dt=target_start_dt,
            target_end_dt=target_end_dt,
        )

    @staticmethod
    def _is_no_changes_move(
        *,
        mode,
        source_schedule,
        target_schedule,
        target_start_dt,
        target_end_dt,
    ):
        if target_schedule is not None and target_schedule.id == source_schedule.id:
            return True

        return (
            mode == "move"
            and source_schedule.start_time == target_start_dt
            and source_schedule.end_time == target_end_dt
        )

    @staticmethod
    def _build_move_assignments(
        *,
        source_schedule,
        target_schedule,
        target_start_dt,
        target_end_dt,
    ):
        original_source_times = (source_schedule.start_time, source_schedule.end_time)
        assignments = {source_schedule.id: (target_start_dt, target_end_dt)}
        changed_ids = {source_schedule.id}
        original_target_times = None

        if target_schedule is not None:
            original_target_times = (
                target_schedule.start_time,
                target_schedule.end_time,
            )
            assignments[target_schedule.id] = original_source_times
            changed_ids.add(target_schedule.id)

        return assignments, changed_ids, original_source_times, original_target_times

    def move(self, request):
        actor = getattr(request.user, "email", "")
        active_team = self.get_active_team()

        parsed_request, parsed_request_error = self._parse_move_request_payload(
            request.data
        )
        if parsed_request_error is not None:
            return parsed_request_error

        mode = parsed_request["mode"]
        source_slot = parsed_request["source_slot"]
        target_slot = parsed_request["target_slot"]

        source_schedule, source_schedule_error = self._fetch_source_schedule_for_move(
            request_user=request.user,
            active_team=active_team,
            source_slot=source_slot,
        )
        if source_schedule_error is not None:
            return source_schedule_error

        target_start_dt, target_end_dt = self._resolve_slot_datetimes_for_source_week(
            source_start=source_schedule.start_time,
            day_name=target_slot["day"],
            start_time=target_slot["start_time"],
            end_time=target_slot["end_time"],
        )

        (
            scope_queryset,
            scope_schedules,
            scope_by_id,
            scope_error,
        ) = self._resolve_move_scope(
            request_user=request.user,
            source_schedule=source_schedule,
            active_team=active_team,
        )
        if scope_error is not None:
            return scope_error

        target_schedule, target_schedule_error = self._resolve_target_schedule_for_mode(
            mode=mode,
            scope_queryset=scope_queryset,
            scope_by_id=scope_by_id,
            source_schedule=source_schedule,
            target_slot=target_slot,
            target_start_dt=target_start_dt,
            target_end_dt=target_end_dt,
        )
        if target_schedule_error is not None:
            return target_schedule_error

        if self._is_no_changes_move(
            mode=mode,
            source_schedule=source_schedule,
            target_schedule=target_schedule,
            target_start_dt=target_start_dt,
            target_end_dt=target_end_dt,
        ):
            return self._move_no_changes_response(mode)

        (
            assignments,
            changed_ids,
            original_source_times,
            original_target_times,
        ) = self._build_move_assignments(
            source_schedule=source_schedule,
            target_schedule=target_schedule,
            target_start_dt=target_start_dt,
            target_end_dt=target_end_dt,
        )

        validation_error = self._validate_minimal_move_constraints(
            scope_schedules=scope_schedules,
            assignments=assignments,
            changed_ids=changed_ids,
        )
        if validation_error is not None:
            return validation_error

        affected_schedules = self._apply_move_assignments(
            source_schedule=source_schedule,
            target_schedule=target_schedule,
            target_start_dt=target_start_dt,
            target_end_dt=target_end_dt,
            original_source_times=original_source_times,
            actor=actor,
        )

        unique_affected_slots = self._build_affected_slot_descriptors(
            original_source_times=original_source_times,
            target_start_dt=target_start_dt,
            target_end_dt=target_end_dt,
            original_target_times=original_target_times,
        )

        serialized = self.get_serializer(affected_schedules, many=True)
        return Response(
            {
                "detail": "Schedule change applied successfully.",
                "mode": "swap" if target_schedule is not None else "move",
                "no_changes": False,
                "affected_schedules": serialized.data,
                "affected_slots": unique_affected_slots,
                "teacher_workloads": self._build_teacher_workloads(scope_schedules),
            },
            status=status.HTTP_200_OK,
        )

    def _parse_analyze_params(self, request):
        """
        Parsea y valida los parámetros del request de análisis.

        Args:
            request: HTTP request object

        Returns:
            Tuple[List[int], str]: (schedule_ids, source)

        Raises:
            ValidationAppError: Si parámetros inválidos
        """
        schedule_ids = request.data.get("schedule_ids", [])
        source = (request.data.get("source") or "").strip().lower()

        if not (schedule_ids or source in {"generated", "saved"}):
            raise ValidationAppError(
                "INVALID_ANALYZE_PARAMS",
                "Se debe especificar schedule_ids o source (generated/saved).",
            )

        return schedule_ids, source

    def _get_schedules_to_analyze(self, queryset, schedule_ids, source):
        """
        Obtiene schedules a analizar filtrando por IDs o fuente.

        Args:
            queryset: Schedule queryset base
            schedule_ids: List de IDs específicos (puede estar vacío)
            source: "generated", "saved", o vacío

        Returns:
            List[Schedule]: Schedules a analizar

        Raises:
            ValidationAppError: Si no hay schedules para analizar
        """
        if schedule_ids and isinstance(schedule_ids, list):
            schedules = list(queryset.filter(id__in=schedule_ids))
        elif source in {"generated", "saved"}:
            schedules = list(self._resolve_source_queryset(queryset, source))
        else:
            schedules = []

        if not schedules:
            raise ValidationAppError(
                "NO_SCHEDULES_FOUND",
                "No se encontraron horarios para analizar.",
            )

        return schedules

    def _parse_and_validate_analysis_request(self, request):
        """
        Parsea, valida y obtiene los schedules del request.

        Args:
            request: HTTP request object

        Returns:
            List[Schedule]: Schedules a analizar

        Raises:
            ValidationAppError: Si hay errores de validación
        """
        schedule_ids, source = self._parse_analyze_params(request)
        queryset = self.get_queryset()
        return self._get_schedules_to_analyze(queryset, schedule_ids, source)

    def _perform_defect_analysis(self, schedules):
        """
        Ejecuta el análisis de defectos en los schedules.

        Args:
            schedules: List[Schedule] a analizar

        Returns:
            List de defectos encontrados

        Raises:
            ValidationAppError: Si hay errores en el análisis
        """
        try:
            return ScheduleEvaluator.analyze_schedules(schedules)
        except Exception as e:
            logger.exception("Error analyzing schedules: %s", str(e))
            raise ValidationAppError(
                "ANALYSIS_ERROR",
                f"Error al analizar el horario: {str(e)}",
            )

    @action(detail=False, methods=["post"], url_path="analyze")
    def analyze(self, request):
        """
        Analiza un conjunto de schedules en busca de defectos.

        Espera: {
            "schedule_ids": [1, 2, 3, ...] o
            "source": "generated" | "saved"
        }

        Retorna:
        {
            "defects": [...]
        }
        """
        schedules = self._parse_and_validate_analysis_request(request)
        defects = self._perform_defect_analysis(schedules)

        return Response(
            {"count": len(defects), "defects": defects},
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["post"], url_path="apply-manual-change")
    def apply_manual_change(self, request):
        """Apply a manual session-to-slot change and replan the entire schedule."""
        actor = getattr(request.user, "email", "")

        schedule_id = request.data.get("schedule_id")
        new_slot_index = request.data.get("new_slot_index")

        # Validate inputs
        if schedule_id is None:
            return Response(
                {"detail": "schedule_id is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if new_slot_index is None:
            return Response(
                {"detail": "new_slot_index is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            schedule_id = int(schedule_id)
            new_slot_index = int(new_slot_index)
        except (TypeError, ValueError):
            return Response(
                {"detail": "schedule_id and new_slot_index must be integers."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if schedule_id <= 0:
            return Response(
                {"schedule_id": "schedule_id must be a positive integer."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if new_slot_index < 0:
            return Response(
                {"new_slot_index": "new_slot_index must be zero or greater."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            new_schedules = ScheduleReplanner.replan_with_manual_change(
                user=request.user,
                team=self.get_active_team(),
                schedule_to_move_id=schedule_id,
                new_slot_index=new_slot_index,
                actor_email=actor,
            )
        except ScheduleGenerationError:
            logger.warning(
                "ScheduleGenerationError while applying manual change: "
                "schedule_id=%s, new_slot_index=%s, actor=%s",
                schedule_id,
                new_slot_index,
                actor,
            )
            raise

        serialized = self.get_serializer(new_schedules, many=True)
        return Response(
            {
                "detail": "Schedule replanned with manual change successfully.",
                "schedules": serialized.data,
                "generated_count": len(serialized.data),
                "teacher_workloads": self._build_teacher_workloads(new_schedules),
            },
            status=status.HTTP_200_OK,
        )
