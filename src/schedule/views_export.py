"""Export helpers for the schedule app (CSV, Excel, PDF).

All functions are pure utilities that build HTTP responses or data structures
from a queryset/unit list.  They are imported by ScheduleViewSet in views.py.
"""

import re
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.response import Response

from common.export_utils import build_csv_response, sanitize_filename_stem

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Column order used for every export entity type.
EXPORT_ENTITY_ORDER = ["teacher", "classroom", "group"]

_DAY_ORDER = {"Lunes": 0, "Martes": 1, "Miércoles": 2, "Jueves": 3, "Viernes": 4}
_TC_DAY_NAMES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]


def build_teacher_workloads(schedules):
    """Compute per-teacher total workload in minutes and hours from a schedule list.
    Input: schedules - iterable of Schedule instances with teacher, start_time, end_time
    Output: list of dicts sorted by teacher name, each with
            teacher_id, teacher_name, total_minutes, total_hours
    """
    workloads = {}

    for schedule in schedules or []:
        teacher = getattr(schedule, "teacher", None)
        start_time = getattr(schedule, "start_time", None)
        end_time = getattr(schedule, "end_time", None)

        if teacher is None or start_time is None or end_time is None:
            continue

        duration_seconds = (end_time - start_time).total_seconds()
        if duration_seconds <= 0:
            continue

        duration_minutes = int(round(duration_seconds / 60.0))
        if duration_minutes <= 0:
            continue

        teacher_name = (getattr(teacher, "name", "") or "").strip()
        if not teacher_name:
            teacher_name = f"Profesor {teacher.id}"

        item = workloads.setdefault(
            teacher.id,
            {
                "teacher_id": teacher.id,
                "teacher_name": teacher_name,
                "total_minutes": 0,
            },
        )
        item["total_minutes"] += duration_minutes

    return [
        {
            **item,
            "total_hours": round(item["total_minutes"] / 60.0, 2),
        }
        for item in sorted(
            workloads.values(),
            key=lambda value: value["teacher_name"].lower(),
        )
    ]


def build_export_rows(schedules):
    """Convert a queryset of schedules to a list of plain row dicts for CSV/table output.
    Input: schedules - iterable of Schedule instances with localised time support
    Output: list of dicts with keys: day, start, end, subject, teacher, group, classroom
    """
    weekday_to_name = {
        0: "Lunes",
        1: "Martes",
        2: "Miércoles",
        3: "Jueves",
        4: "Viernes",
        5: "Sábado",
        6: "Domingo",
    }

    rows = []
    for schedule in schedules:
        local_start = (
            timezone.localtime(schedule.start_time) if schedule.start_time else None
        )
        local_end = timezone.localtime(schedule.end_time) if schedule.end_time else None
        rows.append(
            {
                "day": (
                    weekday_to_name.get(local_start.weekday(), "")
                    if local_start
                    else ""
                ),
                "start": local_start.strftime("%H:%M") if local_start else "",
                "end": local_end.strftime("%H:%M") if local_end else "",
                "subject": schedule.subject.name if schedule.subject else "",
                "teacher": schedule.teacher.name if schedule.teacher else "",
                "group": schedule.group.name if schedule.group else "",
                "classroom": schedule.classroom.name if schedule.classroom else "",
            }
        )
    return rows


def build_tc_export_rows(tc_sessions):
    """Convert TCSession instances to row dicts compatible with CSV/Excel export.
    Input: tc_sessions - iterable of TCSession instances with teacher, day, start_time, end_time
    Output: list of dicts with keys: day, start, end, subject, teacher, group, classroom
    """
    rows = []
    for tc in tc_sessions:
        day_name = _TC_DAY_NAMES[tc.day] if 0 <= tc.day <= 4 else ""
        teacher_name = (tc.teacher.name if tc.teacher else "") or ""
        rows.append(
            {
                "day": day_name,
                "start": tc.start_time.strftime("%H:%M") if tc.start_time else "",
                "end": tc.end_time.strftime("%H:%M") if tc.end_time else "",
                "subject": "Guardia",
                "teacher": teacher_name,
                "group": "",
                "classroom": "",
            }
        )
    return rows


def collect_tc_slots_and_content(tc_sessions, show_teacher_names=False):
    """Collect TC session time slots and cell content for timetable grid rendering.
    Input: tc_sessions - iterable of TCSession instances;
           show_teacher_names - if True cell content is the teacher name, else 'Guardia'
    Output: tuple (slot_keys, cell_content) where slot_keys is a list of (start, end)
            tuples and cell_content is a dict {(day_name, slot): [content_str, ...]}
    """
    slot_keys = []
    cell_content = {}
    for tc in tc_sessions:
        if not (0 <= tc.day <= 4):
            continue
        day_name = _TC_DAY_NAMES[tc.day]
        slot = (
            tc.start_time.strftime("%H:%M") if tc.start_time else "",
            tc.end_time.strftime("%H:%M") if tc.end_time else "",
        )
        if slot not in slot_keys:
            slot_keys.append(slot)
        key = (day_name, slot)
        content = (
            (tc.teacher.name if tc.teacher else "?")
            if show_teacher_names
            else "Guardia"
        )
        entry_list = cell_content.setdefault(key, [])
        if content not in entry_list:
            entry_list.append(content)
    return slot_keys, cell_content


def _build_tc_roster_unit(tc_sessions):
    """Build an export unit representing the full duty-hour roster.
    Input: tc_sessions - list of TCSession instances
    Output: unit dict with entity_type 'tc_roster'
    """
    rows = build_tc_export_rows(tc_sessions)
    return {
        "entity_type": "tc_roster",
        "header": "Horas de guardia",
        "rows": rows,
        "schedules": [],
        "tc_sessions": list(tc_sessions),
    }


def _build_entity_name_and_stage_maps(entity_type, model_cls, selected_ids, active_team):
    if entity_type == "group":
        group_objs = list(
            model_cls.objects.filter(
                id__in=selected_ids,
                team=active_team,
            ).only("id", "name", "stage")
        )
        name_map = {obj.id: obj.name for obj in group_objs}
        stage_map = {obj.id: getattr(obj, "stage", "") for obj in group_objs}
    else:
        name_map = {
            obj.id: obj.name
            for obj in model_cls.objects.filter(
                id__in=selected_ids,
                team=active_team,
            ).only("id", "name")
        }
        stage_map = {}
    return name_map, stage_map


def _build_card_export_units(queryset, filters, export_entity_config, active_team, tc_by_teacher_id):
    units = []
    for entity_type in EXPORT_ENTITY_ORDER:
        config = export_entity_config[entity_type]
        field_name = config["field"]
        entity_filter = filters[entity_type]

        selected_ids = []
        if entity_filter["include_all"]:
            selected_ids.extend(
                list(
                    queryset.exclude(**{f"{field_name}_id__isnull": True})
                    .values_list(f"{field_name}_id", flat=True)
                    .distinct()
                )
            )
        selected_ids.extend(entity_filter["ids"])
        selected_ids = sorted(set(selected_ids))
        if not selected_ids:
            continue

        name_map, stage_map = _build_entity_name_and_stage_maps(
            entity_type, config["model"], selected_ids, active_team
        )

        for object_id in selected_ids:
            object_name = name_map.get(object_id, f"{config['label']} {object_id}")
            object_queryset = (
                queryset.filter(**{f"{field_name}_id": object_id})
                .order_by("start_time", "id")
                .distinct()
            )
            rows = build_export_rows(object_queryset)
            unit_tc = (
                tc_by_teacher_id.get(object_id, []) if entity_type == "teacher" else []
            )
            if unit_tc:
                tc_rows = build_tc_export_rows(unit_tc)
                rows = sorted(
                    rows + tc_rows,
                    key=lambda r: (_DAY_ORDER.get(r["day"], 99), r["start"]),
                )
            if entity_type == "group":
                rows = _inject_group_recess_rows(
                    rows, stage_map.get(object_id, ""), object_name
                )
            units.append(
                {
                    "entity_type": entity_type,
                    "header": f"{config['label']} {object_name}",
                    "rows": rows,
                    "schedules": list(object_queryset),
                    "tc_sessions": unit_tc,
                }
            )
    return units


def build_export_units(
    queryset,
    params,
    active_team,
    export_entity_config,
    tc_sessions=None,
    add_tc_roster=True,
):
    """Build the list of export units (one per card/entity or one global unit).
    Input: queryset - filtered Schedule queryset;
           params - parsed export params dict;
           active_team - Team instance;
           export_entity_config - dict of entity configs from ScheduleViewSet;
           tc_sessions - optional list of TCSession instances to include;
           add_tc_roster - if True append a TC roster summary unit at the end
    Output: list of unit dicts with entity_type, header, rows, schedules, tc_sessions
    """
    tc_by_teacher_id = {}
    if tc_sessions:
        for tc in tc_sessions:
            tc_by_teacher_id.setdefault(tc.teacher_id, []).append(tc)

    units = []
    if params.get("scope") != "cards":
        rows = build_export_rows(queryset)
        if tc_sessions:
            tc_rows = build_tc_export_rows(tc_sessions)
            rows = sorted(
                rows + tc_rows,
                key=lambda r: (_DAY_ORDER.get(r["day"], 99), r["start"]),
            )
        units.append(
            {
                "entity_type": "mixed",
                "header": "",
                "rows": rows,
                "schedules": list(queryset),
                "tc_sessions": list(tc_sessions) if tc_sessions else [],
            }
        )
        if tc_sessions and add_tc_roster:
            units.append(_build_tc_roster_unit(tc_sessions))
        return units

    filters = params["card_filters"]["filters"]
    units = _build_card_export_units(
        queryset, filters, export_entity_config, active_team, tc_by_teacher_id
    )

    if tc_sessions and add_tc_roster:
        units.append(_build_tc_roster_unit(tc_sessions))

    return units


def resolve_saved_schedule_name(params, queryset):
    """Determine the display name for a saved timetable export.
    Input: params - parsed export params dict; queryset - filtered Schedule queryset
    Output: name string from params, first distinct schedule name, or empty string
    """
    explicit_name = (params.get("saved_timetable_name") or "").strip()
    if explicit_name:
        return explicit_name

    if params.get("source") != "saved":
        return ""

    names = list(
        queryset.exclude(name__isnull=True)
        .exclude(name__exact="")
        .values_list("name", flat=True)
        .distinct()[:1]
    )
    if names:
        return names[0]
    return ""


def build_export_filename(params, saved_schedule_name=""):
    """Build the export download filename from params and optional saved name.
    Input: params - parsed export params dict with 'source' and 'format';
           saved_schedule_name - display name of the saved timetable, if any
    Output: filename string including extension (e.g. 'my_schedule.csv')
    """
    if params["source"] == "saved" and saved_schedule_name:
        stem = sanitize_filename_stem(saved_schedule_name, "orarioo_saved_schedule")
    else:
        date_token = timezone.now().strftime("%Y%m%d_%H%M%S")
        stem = f"orarioo_generated_schedule_{date_token}"
    return f"{stem}.{params['format']}"


def build_csv_response_for_schedule(rows, filename):
    """Build an HTTP response containing schedule data as a CSV attachment.
    Input: rows - list of row dicts from build_export_rows; filename - attachment filename
    Output: HttpResponse with CSV content
    """
    header = [
        "Día",
        "Inicio",
        "Fin",
        "Asignatura",
        "Profesor",
        "Curso",
        "Aula",
    ]
    return build_csv_response(
        header,
        [
            [
                row["day"],
                row["start"],
                row["end"],
                row["subject"],
                row["teacher"],
                row["group"],
                row["classroom"],
            ]
            for row in rows
        ],
        filename,
    )


def make_unique_sheet_title(base_title, used_titles):
    """Generate a unique Excel sheet title within the 31-character limit.
    Input: base_title - desired sheet name; used_titles - mutable set of already used names
    Output: unique string title (max 31 chars); side-effect: adds the title to used_titles
    """
    safe = re.sub(r"[\\/*?:\[\]]+", "_", base_title).strip() or "Horario"
    safe = safe[:31]
    if safe not in used_titles:
        used_titles.add(safe)
        return safe
    counter = 2
    while True:
        suffix = f"_{counter}"
        candidate = f"{safe[:31 - len(suffix)]}{suffix}"
        if candidate not in used_titles:
            used_titles.add(candidate)
            return candidate
        counter += 1


def _populate_excel_sheet(sheet, headers, rows):
    """Write headers and data rows to an openpyxl worksheet.
    Input: sheet - openpyxl Worksheet; headers - list of column header strings;
           rows - list of row dicts from build_export_rows
    Output: None; side-effect: appends rows to sheet
    """
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row["day"],
                row["start"],
                row["end"],
                row["subject"],
                row["teacher"],
                row["group"],
                row["classroom"],
            ]
        )


def build_excel_workbook(units):
    """Create an openpyxl Workbook with one sheet per export unit.
    Input: units - list of unit dicts from build_export_units
    Output: openpyxl Workbook instance
    """
    from openpyxl import Workbook

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    headers = ["Día", "Inicio", "Fin", "Asignatura", "Profesor", "Curso", "Aula"]
    used_titles = set()

    if not units:
        sheet = workbook.create_sheet(title="Sin datos")
        sheet.append(headers)
    else:
        for index, unit in enumerate(units, start=1):
            base_title = unit["header"] or f"Horario {index}"
            sheet_title = make_unique_sheet_title(base_title, used_titles)
            sheet = workbook.create_sheet(title=sheet_title)
            _populate_excel_sheet(sheet, headers, unit["rows"])

    return workbook


def build_excel_response(units, filename):
    """Build an Excel HTTP response with multiple sheets, falling back to CSV if needed.
    Input: units - list of unit dicts; filename - attachment filename (.xlsx)
    Output: HttpResponse with Excel or CSV content
    """
    try:
        workbook = build_excel_workbook(units)
    except ImportError:
        merged_rows = []
        for unit in units:
            merged_rows.extend(unit["rows"])
        fallback_filename = re.sub(r"\.xlsx$", ".csv", filename)
        return build_csv_response_for_schedule(merged_rows, fallback_filename)

    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def normalize_stage(stage_value):
    """Normalise a stage name string to one of the three internal stage codes.
    Input: stage_value - raw stage string (may be Spanish or English, any case)
    Output: 'preschool', 'primary', 'secondary', or the lowercased input if unrecognised
    """
    value = (stage_value or "").strip().lower()
    if "preschool" in value or "infantil" in value:
        return "preschool"
    if "primary" in value or "primaria" in value:
        return "primary"
    if "secondary" in value or "eso" in value:
        return "secondary"
    return value


def _describe_schedule(schedule):
    """Return the subject name for a schedule, or '-' if no subject is set.
    Input: schedule - Schedule instance
    Output: subject name string or '-'
    """
    return schedule.subject.name if schedule.subject else "-"


STAGE_BREAK_SLOTS = {
    "preschool": [("10:30", "11:00"), ("13:30", "14:00")],
    "primary": [("11:30", "12:00")],
    "secondary": [("11:00", "11:30")],
}


def collect_slots_and_content(schedules):
    """Collect all time slots and cell content from a schedule list for timetable rendering.
    Input: schedules - iterable of Schedule instances
    Output: tuple (slot_keys, cell_content, day_stage_map) where
            slot_keys is a list of (start_str, end_str) tuples,
            cell_content is a dict {(day_name, slot): [content_str, ...]},
            day_stage_map is a dict {day_name: set(stage_code)}
    """
    weekday_to_name = {
        0: "Lunes",
        1: "Martes",
        2: "Miércoles",
        3: "Jueves",
        4: "Viernes",
    }
    slot_keys = []
    cell_content = {}
    day_stage_map = {}

    for schedule in schedules:
        local_start = timezone.localtime(schedule.start_time)
        local_end = timezone.localtime(schedule.end_time)
        weekday = local_start.weekday()
        if weekday not in weekday_to_name:
            continue

        day_name = weekday_to_name[weekday]
        slot = (local_start.strftime("%H:%M"), local_end.strftime("%H:%M"))
        if slot not in slot_keys:
            slot_keys.append(slot)

        key = (day_name, slot)
        content = _describe_schedule(schedule)
        entry_list = cell_content.setdefault(key, [])
        if content not in entry_list:
            entry_list.append(content)

        normalized_stage = normalize_stage(
            getattr(schedule.group, "stage", "") if schedule.group else ""
        )
        if normalized_stage in STAGE_BREAK_SLOTS:
            day_stage_map.setdefault(day_name, set()).add(normalized_stage)

    return slot_keys, cell_content, day_stage_map


def inject_recess_breaks(slot_keys, cell_content, day_stage_map):
    """Insert recess slots into the timetable data for group (course) schedule views.
    Input: slot_keys - mutable list of (start, end) slot tuples;
           cell_content - mutable dict {(day, slot): [content, ...]};
           day_stage_map - dict {day_name: set(stage_code)}
    Output: None; side-effect: mutates slot_keys and cell_content in place
    """
    for day_name, stages in day_stage_map.items():
        for stage in stages:
            for recess_slot in STAGE_BREAK_SLOTS.get(stage, []):
                if recess_slot not in slot_keys:
                    slot_keys.append(recess_slot)
                cell_content.setdefault((day_name, recess_slot), []).append("Recreo")


def _inject_group_recess_rows(rows, group_stage, group_name):
    """Append recess time slot rows for a group's stage and re-sort by day and start time.
    Input: rows - list of row dicts from build_export_rows;
           group_stage - raw stage string from Group.stage;
           group_name - display name of the group
    Output: new sorted list including recess rows
    """
    normalized = normalize_stage(group_stage)
    breaks = STAGE_BREAK_SLOTS.get(normalized, [])
    if not breaks:
        return rows
    existing_days = {row["day"] for row in rows if row["day"]}
    recess_rows = [
        {
            "day": day,
            "start": start_hm,
            "end": end_hm,
            "subject": "Recreo",
            "teacher": "",
            "group": group_name,
            "classroom": "",
        }
        for day in existing_days
        for start_hm, end_hm in breaks
    ]
    combined = rows + recess_rows
    combined.sort(key=lambda r: (_DAY_ORDER.get(r["day"], 99), r["start"]))
    return combined


def build_timetable_rows(slot_keys, cell_content):
    """Build table row data from collected slots and cell content.
    Input: slot_keys - list of (start, end) tuples; cell_content - dict {(day, slot): [str]}
    Output: list of rows including a header row; each row is a list of cell strings
    """
    days = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
    table_data = [["Hora", *days]]
    for slot in slot_keys:
        row = [f"{slot[0]} - {slot[1]}"]
        for day in days:
            entries = cell_content.get((day, slot), [])
            row.append("\n\n".join(entries))
        table_data.append(row)
    return table_data


def _merge_tc_into_timetable(slot_keys, cell_content, tc_sessions):
    tc_slot_keys, tc_cell_content = collect_tc_slots_and_content(
        tc_sessions, show_teacher_names=False
    )
    for slot in tc_slot_keys:
        if slot not in slot_keys:
            slot_keys.append(slot)
    for (day, slot), contents in tc_cell_content.items():
        entry_list = cell_content.setdefault((day, slot), [])
        for content in contents:
            if content not in entry_list:
                entry_list.append(content)


def build_timetable_table_data(schedules, entity_type, tc_sessions=None):
    """Build the full table data matrix for a timetable PDF page.
    Input: schedules - list of Schedule instances for one entity;
           entity_type - 'teacher', 'group', 'classroom', 'mixed', or 'tc_roster';
           tc_sessions - optional list of TCSession instances to merge into the grid
    Output: list of rows (including header) ready for a ReportLab Table
    """
    _empty_header = ["Hora", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]

    if entity_type == "tc_roster":
        slot_keys, cell_content = collect_tc_slots_and_content(
            tc_sessions or [], show_teacher_names=True
        )
        slot_keys.sort()
        if not slot_keys:
            return [_empty_header, ["Sin horas de guardia", "", "", "", "", ""]]
        return build_timetable_rows(slot_keys, cell_content)

    slot_keys, cell_content, day_stage_map = collect_slots_and_content(schedules)

    if tc_sessions and entity_type == "teacher":
        _merge_tc_into_timetable(slot_keys, cell_content, tc_sessions)

    if entity_type == "group":
        inject_recess_breaks(slot_keys, cell_content, day_stage_map)

    slot_keys.sort()
    if not slot_keys:
        return [_empty_header, ["Sin sesiones", "", "", "", "", ""]]

    return build_timetable_rows(slot_keys, cell_content)


def build_pdf_units_response(units, filename):
    """Build a PDF HTTP response with one timetable page per export unit.
    Input: units - list of unit dicts from build_export_units; filename - attachment filename
    Output: HttpResponse with PDF content, or 503 Response if reportlab is unavailable
    """
    if not REPORTLAB_AVAILABLE:
        return Response(
            {
                "detail": (
                    "PDF export is unavailable because reportlab is not installed. "
                    "Install reportlab to enable PDF exports."
                )
            },
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    story = []

    for index, unit in enumerate(units):
        if index > 0:
            story.append(PageBreak())

        if unit["header"]:
            story.extend(
                [
                    Paragraph(f"<b>{unit['header']}</b>", styles["Title"]),
                    Spacer(1, 10),
                ]
            )

        table_data = build_timetable_table_data(
            unit["schedules"], unit["entity_type"], unit.get("tc_sessions") or None
        )

        available_width = document.width
        time_col_width = available_width * 0.15
        day_col_width = (available_width - time_col_width) / 5
        col_widths = [time_col_width, *([day_col_width] * 5)]

        table = Table(table_data, repeatRows=1, colWidths=col_widths)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e7ecfb")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#b7bfd4")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)

    document.build(story)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
