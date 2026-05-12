from datetime import datetime, timedelta
from io import BytesIO
from types import SimpleNamespace
from unittest import skipIf

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APITestCase

from auditableEntity.models import AuditActionType, AuditEntry
from classroom.models import Classroom
from common.test_utils import AuthenticatedAdminAPIMixin
from group.models import EducationalStage, Group
from schedule.algorithm import assignment as schedule_assignment
from schedule.algorithm.diagnostics import collect_generation_diagnostics
from schedule.algorithm.generator import BasicScheduleGenerator
from schedule.algorithm.slots import (
    STAGE_PRIMARY,
    build_slot_preference_index,
    build_weekly_slots,
    build_windows_from_stage_config,
    parse_schedule_config_to_slot_windows,
    slot_instance_key,
    slot_preference_key_from_datetime,
)
from schedule.algorithm.tc_assigner import assign_tc_sessions
from schedule.constants import AUTO_GENERATED_OBSERVATION, SAVED_TIMETABLE_PREFIX
from schedule.models import Schedule, TCSession
from subject.models import Subject, SubjectTimePreferenceState, SubjectType
from teacher.models import Teacher, TeacherTimePreferenceState

try:
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ScheduleApiTests(AuthenticatedAdminAPIMixin, APITestCase):
    def setUp(self):
        self.authenticate_admin(email_prefix="schedule-api")
        self.other_user = self.create_user(
            email="schedule-api-2@test.com",
            given_name="Api2",
            family_name="Tester2",
        )
        self.team.members.add(self.other_user)

        self.teacher = Teacher.objects.create(
            team=self.team,
            name="Ana Perez",
            max_weekly_hours=20,
            working_hours=12,
        )
        self.classroom = Classroom.objects.create(name="Aula 1A", team=self.team)
        self.group = Group.objects.create(
            name="1A",
            stage=EducationalStage.PRIMARY,
            team=self.team,
        )
        self.subject = Subject.objects.create(
            team=self.team,
            name="Mathematics",
            weekly_hours=5,
            duration=1.5,

            type=SubjectType.NORMAL,
            teacher=self.teacher,
            group=self.group,
        )

    def build_payload(self):
        start_time = timezone.now() + timedelta(days=1)
        end_time = start_time + timedelta(hours=1)
        return {
            "name": "Math Monday",
            "start_time": start_time.isoformat(),
            "end_time": end_time.isoformat(),
            "observations": "Bring calculators",
            "teacher": self.teacher.id,
            "classroom": self.classroom.id,
            "group": self.group.id,
            "subject": self.subject.id,
            "users": [self.user.id],
        }

    def generate_schedule(self, payload=None):
        return self.client.post(
            reverse("schedule-generate"), payload or {}, format="json"
        )

    def assert_generate_bad_request_with_detail(self, response, detail_snippet):
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("detail", response.data)
        self.assertIn("_error", response.data)
        self.assertIn("errors", response.data)
        self.assertEqual(response.data["_meta"]["success"], False)
        candidates = [response.data.get("detail", "")]
        for entry in response.data.get("errors", {}).get("non_field_errors", []):
            candidates.append(str(entry.get("message", "")))
        self.assertTrue(
            any(detail_snippet in text for text in candidates),
            f"'{detail_snippet}' not found in detail or error messages: {candidates}",
        )

    def assert_generate_bad_request_has_codes(self, response, *codes):
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("errors", response.data)
        non_field_errors = response.data["errors"].get("non_field_errors", [])
        returned_codes = [entry.get("code") for entry in non_field_errors]
        for code in codes:
            self.assertIn(code, returned_codes)

    def create_schedule(
        self,
        *,
        name="Science",
        start_time=None,
        end_time=None,
        observations="Lab class",
        created_by="",
        updated_by="",
        teacher=None,
        classroom=None,
        group=None,
        subject=None,
        users=None,
    ):
        start_time = start_time or (timezone.now() + timedelta(days=1))
        end_time = end_time or (start_time + timedelta(hours=1))
        schedule = Schedule.objects.create(
            team=self.team,
            name=name,
            start_time=start_time,
            end_time=end_time,
            observations=observations,
            teacher=teacher or self.teacher,
            classroom=classroom or self.classroom,
            group=group or self.group,
            subject=subject or self.subject,
            created_by=created_by,
            updated_by=updated_by,
        )
        for user in users or [self.user]:
            schedule.users.add(user)
        return schedule

    @staticmethod
    def slot_descriptor_from_datetimes(start_time, end_time):
        weekday_to_name = {
            0: "Lunes",
            1: "Martes",
            2: "Miércoles",
            3: "Jueves",
            4: "Viernes",
        }
        local_start = timezone.localtime(start_time)
        local_end = timezone.localtime(end_time)
        return {
            "day": weekday_to_name[local_start.weekday()],
            "start": local_start.strftime("%H:%M"),
            "end": local_end.strftime("%H:%M"),
        }

    @staticmethod
    def minutes_for_teacher_from_serialized_schedules(items, teacher_id):
        total_minutes = 0
        for item in items or []:
            if int(item.get("teacher") or 0) != int(teacher_id):
                continue

            start_time = item.get("start_time")
            end_time = item.get("end_time")
            if not start_time or not end_time:
                continue

            start_dt = datetime.fromisoformat(start_time.replace("Z", "+00:00"))
            end_dt = datetime.fromisoformat(end_time.replace("Z", "+00:00"))
            duration_minutes = int(round((end_dt - start_dt).total_seconds() / 60))
            if duration_minutes > 0:
                total_minutes += duration_minutes

        return total_minutes

    def test_create_schedule(self):
        response = self.client.post(
            reverse("schedule-list"),
            self.build_payload(),
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Schedule.objects.count(), 1)

    def test_list_and_retrieve_schedule(self):
        schedule = self.create_schedule()

        list_response = self.client.get(reverse("schedule-list"))
        detail_response = self.client.get(
            reverse("schedule-detail", args=[schedule.id])
        )

        self.assertEqual(list_response.status_code, status.HTTP_200_OK)
        self.assertEqual(detail_response.status_code, status.HTTP_200_OK)
        self.assertEqual(detail_response.data["name"], "Science")

    def test_export_csv_includes_global_schedules(self):
        created_schedule = self.create_schedule(
            name="Horario Usuario 1",
            observations=AUTO_GENERATED_OBSERVATION,
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )
        self.create_schedule(
            name="Horario Usuario 2",
            observations=AUTO_GENERATED_OBSERVATION,
            created_by=self.other_user.email,
            updated_by=self.other_user.email,
            users=[self.other_user],
        )

        response = self.client.get(
            reverse("schedule-export"),
            {"export_format": "csv", "source": "generated", "scope": "all"},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        csv_text = response.content.decode("utf-8-sig")
        self.assertIn("Día,Inicio,Fin,Asignatura,Profesor,Curso,Aula", csv_text)
        self.assertIn("Mathematics", csv_text)
        self.assertIn("Ana Perez", csv_text)
        self.assertIn("1A", csv_text)
        self.assertIn("Aula 1A", csv_text)
        self.assertIn(
            timezone.localtime(created_schedule.start_time).strftime("%H:%M"),
            csv_text,
        )
        self.assertIn(
            timezone.localtime(created_schedule.end_time).strftime("%H:%M"),
            csv_text,
        )
        self.assertTrue(
            any(
                day_name in csv_text
                for day_name in [
                    "Lunes",
                    "Martes",
                    "Miércoles",
                    "Jueves",
                    "Viernes",
                    "Sábado",
                    "Domingo",
                ]
            )
        )
        self.assertIn("orarioo_generated_schedule_", response["Content-Disposition"])

    def test_export_csv_by_group_entity_filter(self):
        other_group = Group.objects.create(
            name="2A",
            stage=EducationalStage.PRIMARY,
            team=self.team,
        )
        other_subject = Subject.objects.create(
            team=self.team,
            name="Language 2A",
            weekly_hours=3,
            duration=1.0,

            type=SubjectType.NORMAL,
            teacher=self.teacher,
            group=other_group,
        )

        self.create_schedule(
            name="Sesion 1A",
            group=self.group,
            subject=self.subject,
            observations=AUTO_GENERATED_OBSERVATION,
        )
        self.create_schedule(
            name="Sesion 2A",
            group=other_group,
            subject=other_subject,
            observations=AUTO_GENERATED_OBSERVATION,
        )

        response = self.client.get(
            reverse("schedule-export"),
            {
                "export_format": "csv",
                "source": "generated",
                "scope": "entity",
                "entity_type": "group",
                "entity_id": self.group.id,
            },
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        csv_text = response.content.decode("utf-8-sig")
        self.assertIn("Mathematics", csv_text)
        self.assertNotIn("Language 2A", csv_text)

    @skipIf(not OPENPYXL_AVAILABLE, "openpyxl is not installed")
    def test_export_cards_mode_with_specific_teacher_without_teacher_all(self):
        second_teacher = Teacher.objects.create(
            team=self.team,
            name="Julian",
            max_weekly_hours=20,
            working_hours=12,
        )
        second_subject = Subject.objects.create(
            team=self.team,
            name="Science",
            weekly_hours=3,
            duration=1.0,

            type=SubjectType.NORMAL,
            teacher=second_teacher,
            group=self.group,
        )

        self.create_schedule(
            name="Sesion Ana",
            teacher=self.teacher,
            subject=self.subject,
            observations=AUTO_GENERATED_OBSERVATION,
        )
        self.create_schedule(
            name="Sesion Julian",
            teacher=second_teacher,
            subject=second_subject,
            observations=AUTO_GENERATED_OBSERVATION,
        )

        response = self.client.get(
            reverse("schedule-export"),
            {
                "export_format": "csv",
                "source": "generated",
                "selection_mode": "cards",
                "group_all": "0",
                "teacher_all": "0",
                "classroom_all": "0",
                "subject_all": "0",
                "teacher_ids": str(second_teacher.id),
            },
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(filename=BytesIO(response.content))
        self.assertGreaterEqual(len(workbook.sheetnames), 1)
        first_sheet = workbook[workbook.sheetnames[0]]
        values = [
            str(value)
            for row in first_sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        ]
        joined = " ".join(values)
        self.assertIn("Julian", joined)
        self.assertNotIn("Ana Perez", joined)

    def test_update_schedule(self):
        schedule = self.create_schedule()
        new_end = timezone.now() + timedelta(days=2, hours=2)

        payload = {
            "observations": "Updated class note",
            "end_time": new_end.isoformat(),
            "users": [self.user.id, self.other_user.id],
        }

        response = self.client.patch(
            reverse("schedule-detail", args=[schedule.id]),
            payload,
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        schedule.refresh_from_db()
        self.assertEqual(schedule.observations, "Updated class note")
        self.assertEqual(schedule.users.count(), 2)

    def test_delete_schedule(self):
        schedule = self.create_schedule()

        response = self.client.delete(reverse("schedule-detail", args=[schedule.id]))

        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)
        self.assertFalse(Schedule.objects.filter(id=schedule.id).exists())

    def test_reject_if_end_time_is_not_greater_than_start_time(self):
        start_time = timezone.now() + timedelta(days=1)
        payload = self.build_payload()
        payload["start_time"] = start_time.isoformat()
        payload["end_time"] = start_time.isoformat()

        response = self.client.post(reverse("schedule-list"), payload, format="json")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("end_time", response.data)

    def test_reject_whitespace_only_name(self):
        payload = self.build_payload()
        payload["name"] = "    "

        response = self.client.post(reverse("schedule-list"), payload, format="json")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("name", response.data)

    def test_reject_empty_users(self):
        payload = self.build_payload()
        payload["users"] = []

        response = self.client.post(reverse("schedule-list"), payload, format="json")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("users", response.data)

    def test_reject_missing_subject(self):
        payload = self.build_payload()
        payload.pop("subject")

        response = self.client.post(reverse("schedule-list"), payload, format="json")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("subject", response.data)

    def test_generate_basic_schedule(self):
        Classroom.objects.all().delete()
        AuditEntry.objects.all().delete()

        response = self.generate_schedule()

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertIn("schedules", response.data)
        self.assertEqual(response.data["generated_count"], self.subject.weekly_hours)
        self.assertIn("teacher_workloads", response.data)

        teacher_workload = next(
            (
                item
                for item in response.data["teacher_workloads"]
                if item.get("teacher_id") == self.teacher.id
            ),
            None,
        )
        self.assertIsNotNone(teacher_workload)
        expected_minutes = self.minutes_for_teacher_from_serialized_schedules(
            response.data["schedules"],
            self.teacher.id,
        )
        self.assertEqual(teacher_workload["total_minutes"], expected_minutes)
        self.assertAlmostEqual(
            teacher_workload["total_hours"],
            expected_minutes / 60,
            places=2,
        )

        self.assertEqual(Schedule.objects.count(), self.subject.weekly_hours)

        schedule = Schedule.objects.first()
        self.assertEqual(schedule.teacher_id, self.teacher.id)
        self.assertEqual(schedule.group_id, self.group.id)
        self.assertEqual(schedule.users.count(), 1)
        self.assertEqual(schedule.users.first().id, self.user.id)
        self.assertIsNotNone(schedule.classroom_id)
        self.assertIsNotNone(schedule.group_id)
        entries = list(
            AuditEntry.objects.filter(
                entity_type="schedule",
                action_type=AuditActionType.CREATE,
            )
        )
        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0].entity_name, "Generacion automatica")
        self.assertEqual(
            entries[0].changed_fields,
            [{"campo": "Sesiones generadas", "valor_nuevo": self.subject.weekly_hours}],
        )

    def test_generate_basic_schedule_avoids_teacher_overlap(self):
        Subject.objects.create(
            team=self.team,
            name="Physics",
            weekly_hours=1,
            duration=1.0,

            type=SubjectType.NORMAL,
            teacher=self.teacher,
            group=self.group,
        )

        response = self.generate_schedule()

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        teacher_schedules = list(
            Schedule.objects.filter(teacher=self.teacher).order_by("start_time")
        )
        unique_starts = {item.start_time for item in teacher_schedules}
        self.assertEqual(len(unique_starts), len(teacher_schedules))

    def test_generate_basic_schedule_avoids_teacher_overlap_across_groups(self):
        group_2 = Group.objects.create(
            name="2A",
            stage=EducationalStage.PRIMARY,
            team=self.team,
        )
        Classroom.objects.create(name="Aula 2A", team=self.team)
        Subject.objects.create(
            team=self.team,
            name="Mathematics 2A",
            weekly_hours=5,
            duration=1.0,

            type=SubjectType.NORMAL,
            teacher=self.teacher,
            group=group_2,
        )

        response = self.generate_schedule()
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        teacher_schedules = list(
            Schedule.objects.filter(teacher=self.teacher).order_by("start_time")
        )
        unique_starts = {item.start_time for item in teacher_schedules}
        self.assertEqual(len(unique_starts), len(teacher_schedules))

    def test_generate_basic_schedule_avoids_group_overlap(self):
        teacher_2 = Teacher.objects.create(
            team=self.team,
            name="Carlos Torres",
            max_weekly_hours=20,
            working_hours=8,
        )
        Subject.objects.create(
            team=self.team,
            name="Science",
            weekly_hours=5,
            duration=1.0,

            type=SubjectType.NORMAL,
            teacher=teacher_2,
            group=self.group,
        )

        response = self.generate_schedule()
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        group_schedules = list(
            Schedule.objects.filter(group=self.group).order_by("start_time")
        )
        unique_starts = {item.start_time for item in group_schedules}
        self.assertEqual(len(unique_starts), len(group_schedules))

    def test_apply_manual_change_replans_saved_timetable(self):
        slots = build_weekly_slots()
        primary_slots = [slot for slot in slots if slot.get("stage") == "PRIMARY"]
        self.assertGreaterEqual(len(primary_slots), 3)

        saved_observation = "Saved timetable: Horario Manual Test"
        schedule_to_move = self.create_schedule(
            name="Horario Manual Test",
            start_time=primary_slots[0]["start"],
            end_time=primary_slots[0]["end"],
            observations=saved_observation,
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )
        self.create_schedule(
            name="Horario Manual Test",
            start_time=primary_slots[1]["start"],
            end_time=primary_slots[1]["end"],
            observations=saved_observation,
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )
        AuditEntry.objects.all().delete()

        target_slot_index = slots.index(primary_slots[2])

        response = self.client.post(
            reverse("schedule-apply-manual-change"),
            {
                "schedule_id": schedule_to_move.id,
                "new_slot_index": target_slot_index,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("schedules", response.data)

        schedule_to_move.refresh_from_db()
        self.assertEqual(schedule_to_move.start_time, primary_slots[2]["start"])
        self.assertEqual(schedule_to_move.end_time, primary_slots[2]["end"])
        self.assertGreaterEqual(
            AuditEntry.objects.filter(
                entity_type="schedule",
                action_type=AuditActionType.UPDATE,
            ).count(),
            1,
        )

    def test_move_endpoint_moves_single_generated_session(self):
        slots = build_weekly_slots()
        primary_slots = [slot for slot in slots if slot.get("stage") == "PRIMARY"]
        self.assertGreaterEqual(len(primary_slots), 3)

        source_schedule = self.create_schedule(
            name="Auto Move",
            start_time=primary_slots[0]["start"],
            end_time=primary_slots[0]["end"],
            observations=AUTO_GENERATED_OBSERVATION,
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )

        source_slot = self.slot_descriptor_from_datetimes(
            source_schedule.start_time,
            source_schedule.end_time,
        )
        target_slot = self.slot_descriptor_from_datetimes(
            primary_slots[2]["start"],
            primary_slots[2]["end"],
        )

        response = self.client.post(
            reverse("schedule-move"),
            {
                "mode": "move",
                "source_slot": {
                    "schedule_id": source_schedule.id,
                    **source_slot,
                },
                "target_slot": target_slot,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["mode"], "move")
        self.assertFalse(response.data["no_changes"])
        self.assertIn("teacher_workloads", response.data)
        self.assertTrue(
            any(
                item.get("teacher_id") == self.teacher.id
                for item in response.data["teacher_workloads"]
            )
        )

        source_schedule.refresh_from_db()
        self.assertEqual(source_schedule.start_time, primary_slots[2]["start"])
        self.assertEqual(source_schedule.end_time, primary_slots[2]["end"])

    def test_move_endpoint_swaps_sessions_when_target_provided(self):
        slots = build_weekly_slots()
        primary_slots = [slot for slot in slots if slot.get("stage") == "PRIMARY"]
        self.assertGreaterEqual(len(primary_slots), 2)

        other_teacher = Teacher.objects.create(
            team=self.team,
            name="Lucia Martin",
            max_weekly_hours=20,
            working_hours=12,
        )
        other_classroom = Classroom.objects.create(name="Aula 2A", team=self.team)
        other_group = Group.objects.create(
            name="2A",
            stage=EducationalStage.PRIMARY,
            team=self.team,
        )
        other_subject = Subject.objects.create(
            team=self.team,
            name="Lengua 2A",
            weekly_hours=2,
            duration=1.0,

            type=SubjectType.NORMAL,
            teacher=other_teacher,
            group=other_group,
        )

        source_schedule = self.create_schedule(
            name="Auto Source",
            start_time=primary_slots[0]["start"],
            end_time=primary_slots[0]["end"],
            observations=AUTO_GENERATED_OBSERVATION,
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )
        target_schedule = self.create_schedule(
            name="Auto Target",
            start_time=primary_slots[1]["start"],
            end_time=primary_slots[1]["end"],
            observations=AUTO_GENERATED_OBSERVATION,
            created_by=self.user.email,
            updated_by=self.user.email,
            teacher=other_teacher,
            classroom=other_classroom,
            group=other_group,
            subject=other_subject,
            users=[self.user],
        )

        source_slot = self.slot_descriptor_from_datetimes(
            source_schedule.start_time,
            source_schedule.end_time,
        )
        target_slot = self.slot_descriptor_from_datetimes(
            target_schedule.start_time,
            target_schedule.end_time,
        )

        response = self.client.post(
            reverse("schedule-move"),
            {
                "mode": "swap",
                "source_slot": {
                    "schedule_id": source_schedule.id,
                    **source_slot,
                },
                "target_slot": {
                    "schedule_id": target_schedule.id,
                    **target_slot,
                },
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["mode"], "swap")

        source_schedule.refresh_from_db()
        target_schedule.refresh_from_db()
        self.assertEqual(source_schedule.start_time, primary_slots[1]["start"])
        self.assertEqual(source_schedule.end_time, primary_slots[1]["end"])
        self.assertEqual(target_schedule.start_time, primary_slots[0]["start"])
        self.assertEqual(target_schedule.end_time, primary_slots[0]["end"])

    def test_move_endpoint_rejects_teacher_overlap_conflict(self):
        slots = build_weekly_slots()
        primary_slots = [slot for slot in slots if slot.get("stage") == "PRIMARY"]
        self.assertGreaterEqual(len(primary_slots), 3)

        other_group = Group.objects.create(
            name="3A",
            stage=EducationalStage.PRIMARY,
            team=self.team,
        )
        other_subject = Subject.objects.create(
            team=self.team,
            name="Ciencias 3A",
            weekly_hours=2,
            duration=1.0,

            type=SubjectType.NORMAL,
            teacher=self.teacher,
            group=other_group,
        )

        source_schedule = self.create_schedule(
            name="Auto Source Conflict",
            start_time=primary_slots[0]["start"],
            end_time=primary_slots[0]["end"],
            observations=AUTO_GENERATED_OBSERVATION,
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )
        blocker_schedule = self.create_schedule(
            name="Auto Blocker",
            start_time=primary_slots[2]["start"],
            end_time=primary_slots[2]["end"],
            observations=AUTO_GENERATED_OBSERVATION,
            created_by=self.user.email,
            updated_by=self.user.email,
            group=other_group,
            subject=other_subject,
            users=[self.user],
        )

        source_slot = self.slot_descriptor_from_datetimes(
            source_schedule.start_time,
            source_schedule.end_time,
        )
        target_slot = self.slot_descriptor_from_datetimes(
            blocker_schedule.start_time,
            blocker_schedule.end_time,
        )

        response = self.client.post(
            reverse("schedule-move"),
            {
                "mode": "move",
                "source_slot": {
                    "schedule_id": source_schedule.id,
                    **source_slot,
                },
                "target_slot": target_slot,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("detail", response.data)
        self.assertIn("Teacher conflict", response.data["detail"])

        source_schedule.refresh_from_db()
        self.assertEqual(source_schedule.start_time, primary_slots[0]["start"])
        self.assertEqual(source_schedule.end_time, primary_slots[0]["end"])

    def test_move_endpoint_allows_saved_move_with_conflicts_in_other_timetables(self):
        slots = build_weekly_slots()
        primary_slots = [slot for slot in slots if slot.get("stage") == "PRIMARY"]
        self.assertGreaterEqual(len(primary_slots), 3)

        saved_name = "Horario Guardado Scope"
        saved_observation = f"Saved timetable: {saved_name}"
        source_schedule = self.create_schedule(
            name=saved_name,
            start_time=primary_slots[0]["start"],
            end_time=primary_slots[0]["end"],
            observations=saved_observation,
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )

        # Same resources in another saved timetable should not block this move.
        other_saved_observation = "Saved timetable: Horario Distinto"
        blocker_other_timetable = self.create_schedule(
            name="Horario Distinto",
            start_time=primary_slots[2]["start"],
            end_time=primary_slots[2]["end"],
            observations=other_saved_observation,
            created_by=self.user.email,
            updated_by=self.user.email,
            teacher=self.teacher,
            classroom=self.classroom,
            group=self.group,
            subject=self.subject,
            users=[self.user],
        )

        source_slot = self.slot_descriptor_from_datetimes(
            source_schedule.start_time,
            source_schedule.end_time,
        )
        target_slot = self.slot_descriptor_from_datetimes(
            blocker_other_timetable.start_time,
            blocker_other_timetable.end_time,
        )

        response = self.client.post(
            reverse("schedule-move"),
            {
                "mode": "move",
                "source_slot": {
                    "schedule_id": source_schedule.id,
                    **source_slot,
                },
                "target_slot": target_slot,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["mode"], "move")

        source_schedule.refresh_from_db()
        blocker_other_timetable.refresh_from_db()
        self.assertEqual(source_schedule.start_time, primary_slots[2]["start"])
        self.assertEqual(source_schedule.end_time, primary_slots[2]["end"])
        self.assertEqual(blocker_other_timetable.start_time, primary_slots[2]["start"])
        self.assertEqual(blocker_other_timetable.end_time, primary_slots[2]["end"])

    def test_save_generated_rejects_non_auto_generated_sessions(self):
        schedule = self.create_schedule()
        schedule.created_by = self.user.email
        schedule.updated_by = self.user.email
        schedule.save(update_fields=["created_by", "updated_by"])

        response = self.client.post(
            reverse("schedule-save-generated"),
            {
                "timetable_name": "Horario Invalido",
                "schedule_ids": [schedule.id],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("detail", response.data)

    def test_save_generated_rejects_whitespace_only_timetable_name(self):
        start_time = timezone.now() + timedelta(days=1)
        end_time = start_time + timedelta(hours=1)
        schedule = self.create_schedule(
            name="Auto Session 1",
            start_time=start_time,
            end_time=end_time,
            observations=AUTO_GENERATED_OBSERVATION,
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )

        response = self.client.post(
            reverse("schedule-save-generated"),
            {
                "timetable_name": "   ",
                "schedule_ids": [schedule.id],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("timetable_name", response.data)

    def test_save_generated_rejects_duplicate_schedule_ids(self):
        start_time = timezone.now() + timedelta(days=1)
        end_time = start_time + timedelta(hours=1)
        schedule = self.create_schedule(
            name="Auto Session 1",
            start_time=start_time,
            end_time=end_time,
            observations=AUTO_GENERATED_OBSERVATION,
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )

        response = self.client.post(
            reverse("schedule-save-generated"),
            {
                "timetable_name": "Horario Duplicado",
                "schedule_ids": [schedule.id, schedule.id],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("schedule_ids", response.data)

    def test_save_generated_rejects_existing_timetable_name(self):
        start_time = timezone.now() + timedelta(days=1)
        end_time = start_time + timedelta(hours=1)

        self.create_schedule(
            name="Horario Repetido",
            start_time=start_time,
            end_time=end_time,
            observations="Saved timetable: Horario Repetido",
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )

        auto_schedule = self.create_schedule(
            name="Auto Session 1",
            start_time=start_time + timedelta(hours=1),
            end_time=end_time + timedelta(hours=1),
            observations=AUTO_GENERATED_OBSERVATION,
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )

        response = self.client.post(
            reverse("schedule-save-generated"),
            {
                "timetable_name": "Horario Repetido",
                "schedule_ids": [auto_schedule.id],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("timetable_name", response.data)
        auto_schedule.refresh_from_db()
        self.assertEqual(auto_schedule.observations, AUTO_GENERATED_OBSERVATION)

    def test_save_generated_persists_timetable_and_assigns_target_users(self):
        start_time = timezone.now() + timedelta(days=1)
        first_schedule = self.create_schedule(
            name="Auto Session 1",
            start_time=start_time,
            end_time=start_time + timedelta(hours=1),
            observations=AUTO_GENERATED_OBSERVATION,
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )
        second_schedule = self.create_schedule(
            name="Auto Session 2",
            start_time=start_time + timedelta(hours=1),
            end_time=start_time + timedelta(hours=2),
            observations=AUTO_GENERATED_OBSERVATION,
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )

        response = self.client.post(
            reverse("schedule-save-generated"),
            {
                "timetable_name": "Horario Compartido",
                "schedule_ids": [first_schedule.id, second_schedule.id],
                "user_ids": [self.other_user.id],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["saved_count"], 2)
        self.assertEqual(len(response.data["schedules"]), 2)

        expected_observation = f"{SAVED_TIMETABLE_PREFIX}: Horario Compartido"
        for schedule in (first_schedule, second_schedule):
            schedule.refresh_from_db()
            self.assertEqual(schedule.name, "Horario Compartido")
            self.assertEqual(schedule.observations, expected_observation)
            self.assertEqual(schedule.updated_by, self.user.email)
            self.assertEqual(
                set(schedule.users.values_list("id", flat=True)),
                {self.user.id, self.other_user.id},
            )

    def test_saved_endpoint_returns_all_team_saved_schedules(self):
        start_time = timezone.now() + timedelta(days=1)
        end_time = start_time + timedelta(hours=1)
        saved_schedule = self.create_schedule(
            name="Horario Guardado",
            start_time=start_time,
            end_time=end_time,
            observations="Saved timetable: Horario Guardado",
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )

        self.create_schedule(
            name="Auto Session",
            start_time=start_time + timedelta(hours=1),
            end_time=end_time + timedelta(hours=1),
            observations=AUTO_GENERATED_OBSERVATION,
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )

        other_saved_schedule = self.create_schedule(
            name="Horario Otro Usuario",
            start_time=start_time + timedelta(hours=2),
            end_time=end_time + timedelta(hours=2),
            observations="Saved timetable: Horario Otro Usuario",
            created_by=self.other_user.email,
            updated_by=self.other_user.email,
            users=[self.other_user],
        )

        response = self.client.get(reverse("schedule-saved"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 2)
        self.assertEqual(len(response.data["results"]), 2)
        returned_ids = {r["id"] for r in response.data["results"]}
        self.assertIn(saved_schedule.id, returned_ids)
        self.assertIn(other_saved_schedule.id, returned_ids)
        self.assertIn("teacher_workloads", response.data)
        self.assertTrue(
            any(
                item.get("teacher_id") == self.teacher.id
                for item in response.data["teacher_workloads"]
            )
        )

    def test_generate_basic_schedule_requires_teacher(self):
        Teacher.objects.all().delete()

        response = self.generate_schedule()

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("detail", response.data)

    def test_generate_basic_schedule_rejects_teacher_over_max_weekly_hours(self):
        self.teacher.max_weekly_hours = 4
        self.teacher.save(update_fields=["max_weekly_hours"])

        response = self.generate_schedule()

        self.assert_generate_bad_request_with_detail(
            response,
            "exceeds",
        )

    def test_apply_manual_change_rejects_negative_slot_index(self):
        saved_observation = "Saved timetable: Horario Manual Test"
        schedule_to_move = self.create_schedule(
            name="Horario Manual Test",
            observations=saved_observation,
            created_by=self.user.email,
            updated_by=self.user.email,
            users=[self.user],
        )

        response = self.client.post(
            reverse("schedule-apply-manual-change"),
            {
                "schedule_id": schedule_to_move.id,
                "new_slot_index": -1,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("new_slot_index", response.data)

    def test_generate_rejects_teacher_over_max_with_multiple_subjects(self):
        group_2 = Group.objects.create(
            name="2B",
            stage=EducationalStage.PRIMARY,
            team=self.team,
        )
        Subject.objects.create(
            team=self.team,
            name="Physics",
            weekly_hours=3,
            duration=1.0,

            type=SubjectType.NORMAL,
            teacher=self.teacher,
            group=group_2,
        )
        self.teacher.max_weekly_hours = 7
        self.teacher.save(update_fields=["max_weekly_hours"])

        # Same teacher teaches 5h (Math) + 3h (Physics) = 8h, which exceeds 7h.
        response = self.generate_schedule()

        self.assert_generate_bad_request_with_detail(
            response,
            "exceeds",
        )

    def test_generate_rejects_group_over_weekly_capacity_for_primary(self):
        self.teacher.max_weekly_hours = 40
        self.teacher.save(update_fields=["max_weekly_hours"])
        Subject.objects.create(
            team=self.team,
            name="Science",
            weekly_hours=21,
            duration=1.0,

            type=SubjectType.NORMAL,
            teacher=self.teacher,
            group=self.group,
        )

        # 5h (Math) + 21h (Science) = 26h, above PRIMARY weekly limit (25h).
        response = self.generate_schedule()

        self.assert_generate_bad_request_with_detail(
            response,
            "exceeds",
        )

    def test_generate_respects_group_daily_capacity_for_primary(self):
        self.teacher.max_weekly_hours = 30
        self.teacher.save(update_fields=["max_weekly_hours"])
        self.subject.weekly_hours = 25
        self.subject.save(update_fields=["weekly_hours"])

        response = self.generate_schedule()
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        daily_counts = {}
        group_schedules = Schedule.objects.filter(group=self.group)
        for schedule in group_schedules:
            day_key = schedule.start_time.date()
            daily_counts[day_key] = daily_counts.get(day_key, 0) + 1

        self.assertEqual(group_schedules.count(), 25)
        self.assertEqual(len(daily_counts), 5)
        self.assertTrue(all(count <= 5 for count in daily_counts.values()))

    def test_generate_accepts_timeout_minutes_and_returns_it_in_response(self):
        response = self.generate_schedule({"timeout_minutes": 15})

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["generation_options"]["timeout_minutes"], 15)

    def test_generate_rejects_non_integer_timeout_minutes(self):
        response = self.generate_schedule({"timeout_minutes": "invalid"})

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("detail", response.data)
        self.assertIn(
            "timeout_minutes must be an integer value",
            response.data["detail"],
        )

    def test_generate_rejects_non_positive_timeout_minutes(self):
        response = self.generate_schedule({"timeout_minutes": 0})

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("detail", response.data)
        self.assertIn(
            "timeout_minutes must be between 1 and 1440",
            response.data["detail"],
        )

    def test_optimization_timeout_uses_explicit_timeout_minutes(self):
        timeout_seconds = schedule_assignment._resolve_optimization_timeout_seconds(
            generation_options={"timeout_minutes": 15},
        )

        self.assertEqual(timeout_seconds, 900.0)

    def test_optimization_timeout_defaults_to_cap_when_no_timeout_set(self):
        timeout_seconds = schedule_assignment._resolve_optimization_timeout_seconds(
            generation_options={},
        )

        self.assertEqual(
            timeout_seconds, schedule_assignment._UNLIMITED_OPTIMIZATION_CAP_SECONDS
        )

    def test_generate_assigns_only_subject_mandatory_classroom(self):
        self.classroom.is_shared = True
        self.classroom.save(update_fields=["is_shared"])
        assigned = Classroom.objects.create(
            name="Aula Asignada",
            is_shared=True,
            team=self.team,
        )
        self.subject.mandatory_classroom = assigned
        self.subject.save(update_fields=["mandatory_classroom"])

        response = self.generate_schedule()
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        generated_classroom_ids = set(
            Schedule.objects.filter(subject=self.subject).values_list(
                "classroom_id", flat=True
            )
        )
        self.assertEqual(generated_classroom_ids, {assigned.id})

    def test_generate_uses_mandatory_classroom_when_set(self):
        self.classroom.name = "Aula 1A"
        self.classroom.is_shared = False
        self.classroom.save(update_fields=["name", "is_shared"])
        music_room = Classroom.objects.create(
            name="Aula de Musica",
            is_shared=True,
            team=self.team,
        )
        self.subject.mandatory_classroom = music_room
        self.subject.save(update_fields=["mandatory_classroom"])

        response = self.generate_schedule()
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        generated_classroom_ids = set(
            Schedule.objects.filter(subject=self.subject).values_list(
                "classroom_id", flat=True
            )
        )
        self.assertEqual(generated_classroom_ids, {music_room.id})

    def test_generate_uses_any_classroom_when_subject_has_no_restrictions(self):
        self.classroom.is_shared = True
        self.classroom.save(update_fields=["is_shared"])

        response = self.generate_schedule()

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        generated_classroom_ids = set(
            Schedule.objects.filter(subject=self.subject).values_list(
                "classroom_id", flat=True
            )
        )
        self.assertIn(self.classroom.id, generated_classroom_ids)

    def test_generate_spreads_sessions_when_one_allowed_classroom_is_shared(self):
        self.classroom.is_shared = True
        self.classroom.save(update_fields=["is_shared"])
        self.subject.weekly_hours = 1
        self.subject.save(update_fields=["weekly_hours"])

        teacher_2 = Teacher.objects.create(
            team=self.team,
            name="Elena Ruiz",
            max_weekly_hours=20,
            working_hours=8,
        )
        group_2 = Group.objects.create(
            name="2A",
            stage=EducationalStage.PRIMARY,
            team=self.team,
        )
        other_subject = Subject.objects.create(
            team=self.team,
            name="Science",
            weekly_hours=1,
            duration=1.0,

            type=SubjectType.NORMAL,
            teacher=teacher_2,
            group=group_2,
        )
        self.subject.mandatory_classroom = self.classroom
        self.subject.save(update_fields=["mandatory_classroom"])
        other_subject.mandatory_classroom = self.classroom
        other_subject.save(update_fields=["mandatory_classroom"])

        response = self.generate_schedule()
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        generated = list(
            Schedule.objects.filter(subject__in=[self.subject, other_subject]).order_by(
                "start_time"
            )
        )
        self.assertEqual(len(generated), 2)
        self.assertEqual({item.classroom_id for item in generated}, {self.classroom.id})
        self.assertEqual(
            len({item.start_time for item in generated}),
            2,
            "Two sessions that share a single classroom must not overlap.",
        )

    def test_generate_rejects_subject_with_all_slots_marked_unavailable(self):
        all_slot_keys = build_slot_preference_index(slots=build_weekly_slots()).values()
        self.subject.time_preferences = {
            key: SubjectTimePreferenceState.UNAVAILABLE for key in all_slot_keys
        }
        self.subject.save(update_fields=["time_preferences"])

        response = self.generate_schedule()

        self.assert_generate_bad_request_with_detail(
            response,
            "Could not generate a feasible schedule",
        )
        self.assert_generate_bad_request_has_codes(
            response,
            "SUBJECT_NO_AVAILABLE_SLOTS",
        )

    @skipIf(
        schedule_assignment.cp_model is None,
        "Requires OR-Tools CP-SAT to validate soft constraints.",
    )
    def test_generate_prefers_prefer_yes_over_prefer_no_for_subject(self):
        self.subject.weekly_hours = 1
        slot_pref_index = build_slot_preference_index(slots=build_weekly_slots())
        first_slot_key = slot_pref_index[0]
        second_slot_key = slot_pref_index[1]
        self.subject.time_preferences = {
            first_slot_key: SubjectTimePreferenceState.PREFER_NO,
            second_slot_key: SubjectTimePreferenceState.PREFER_YES,
        }
        self.subject.save(update_fields=["weekly_hours", "time_preferences"])

        response = self.generate_schedule()
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        generated = Schedule.objects.get(subject=self.subject)
        generated_slot_key = slot_preference_key_from_datetime(
            slot=generated.start_time
        )
        self.assertEqual(generated_slot_key, second_slot_key)

    def test_generate_rejects_teacher_with_all_slots_marked_unavailable(self):
        all_slot_keys = build_slot_preference_index(slots=build_weekly_slots()).values()
        self.teacher.time_preferences = {
            key: TeacherTimePreferenceState.UNAVAILABLE for key in all_slot_keys
        }
        self.teacher.save(update_fields=["time_preferences"])

        response = self.generate_schedule()

        self.assert_generate_bad_request_with_detail(
            response,
            "Could not generate a feasible schedule",
        )
        self.assert_generate_bad_request_has_codes(
            response,
            "TEACHER_NO_AVAILABLE_SLOTS",
        )

    def test_generate_returns_structured_diagnostics_when_teachers_are_missing(self):
        Teacher.objects.filter(team=self.team).delete()

        response = self.generate_schedule()

        self.assert_generate_bad_request_has_codes(
            response,
            "MISSING_TEACHERS",
            "MISSING_SUBJECTS",
        )

    def test_collect_generation_diagnostics_skips_missing_configuration_when_not_provided(
        self,
    ):
        slots = build_weekly_slots()
        sessions = BasicScheduleGenerator._build_sessions(
            subjects=[self.subject],
            fallback_teacher=self.teacher,
        )

        diagnostics = collect_generation_diagnostics(
            sessions=sessions,
            slots=slots,
            classrooms=[self.classroom],
            generation_options={},
        )

        codes = [entry["code"] for entry in diagnostics]
        self.assertNotIn("MISSING_TEACHERS", codes)
        self.assertNotIn("MISSING_SUBJECTS", codes)

    def test_generate_returns_teacher_insufficient_available_slots_diagnostic(self):
        self.subject.weekly_hours = 2
        self.subject.save(update_fields=["weekly_hours"])
        slot_pref_index = build_slot_preference_index(slots=build_weekly_slots())
        self.teacher.time_preferences = {
            key: TeacherTimePreferenceState.UNAVAILABLE
            for key in slot_pref_index.values()
            if key != "MON_09:00"
        }
        self.teacher.save(update_fields=["time_preferences"])

        response = self.generate_schedule()

        self.assert_generate_bad_request_has_codes(
            response,
            "TEACHER_INSUFFICIENT_AVAILABLE_SLOTS",
            "SUBJECT_INSUFFICIENT_AVAILABLE_SLOTS",
        )

    def test_generate_returns_classroom_bottleneck_diagnostic(self):
        self.subject.weekly_hours = 1
        self.subject.save(update_fields=["weekly_hours"])
        self.subject.mandatory_classroom = self.classroom
        self.subject.save(update_fields=["mandatory_classroom"])

        other_teacher = Teacher.objects.create(
            team=self.team,
            name="Lucia Lopez",
            max_weekly_hours=20,
            working_hours=8,
        )
        other_group = Group.objects.create(
            name="2A",
            stage=EducationalStage.PRIMARY,
            team=self.team,
        )
        other_subject = Subject.objects.create(
            team=self.team,
            name="Science 2A",
            weekly_hours=1,
            duration=1.0,

            type=SubjectType.NORMAL,
            teacher=other_teacher,
            group=other_group,
        )
        other_subject.mandatory_classroom = self.classroom
        other_subject.save(update_fields=["mandatory_classroom"])

        slot_pref_index = build_slot_preference_index(slots=build_weekly_slots())
        self.teacher.time_preferences = {
            key: TeacherTimePreferenceState.UNAVAILABLE
            for key in slot_pref_index.values()
            if key != "MON_09:00"
        }
        other_teacher.time_preferences = dict(self.teacher.time_preferences)
        self.teacher.save(update_fields=["time_preferences"])
        other_teacher.save(update_fields=["time_preferences"])

        response = self.generate_schedule()

        self.assert_generate_bad_request_has_codes(
            response,
            "CLASSROOM_BOTTLENECK",
        )

    def test_generate_returns_multiple_sorted_diagnostics(self):
        self.subject.weekly_hours = 26
        self.subject.save(update_fields=["weekly_hours"])
        self.teacher.max_weekly_hours = 10
        self.teacher.save(update_fields=["max_weekly_hours"])

        response = self.generate_schedule()

        self.assert_generate_bad_request_has_codes(
            response,
            "GROUP_WEEKLY_CAPACITY_EXCEEDED",
            "TEACHER_WEEKLY_CAPACITY_EXCEEDED",
        )
        non_field_errors = response.data["errors"]["non_field_errors"]
        self.assertEqual(non_field_errors[0]["code"], "GROUP_WEEKLY_CAPACITY_EXCEEDED")

    @skipIf(
        schedule_assignment.cp_model is None,
        "Requires OR-Tools CP-SAT to validate soft constraints.",
    )
    def test_generate_prefers_teacher_prefer_yes_over_prefer_no(self):
        self.subject.weekly_hours = 1
        slot_pref_index = build_slot_preference_index(slots=build_weekly_slots())
        first_slot_key = slot_pref_index[0]
        second_slot_key = slot_pref_index[1]
        self.teacher.time_preferences = {
            first_slot_key: TeacherTimePreferenceState.PREFER_NO,
            second_slot_key: TeacherTimePreferenceState.PREFER_YES,
        }
        self.subject.save(update_fields=["weekly_hours"])
        self.teacher.save(update_fields=["time_preferences"])

        response = self.generate_schedule()
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        generated = Schedule.objects.get(subject=self.subject)
        generated_slot_key = slot_preference_key_from_datetime(
            slot=generated.start_time
        )
        self.assertEqual(generated_slot_key, second_slot_key)

    def test_generate_spreads_subject_sessions_across_days(self):
        response = self.generate_schedule()

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        subject_schedules = Schedule.objects.filter(subject=self.subject)
        distinct_days = {s.start_time.date() for s in subject_schedules}
        self.assertGreaterEqual(
            len(distinct_days),
            4,
            "Expected subject sessions to be spread across at least 4 different days.",
        )

    @skipIf(
        schedule_assignment.cp_model is None,
        "Requires OR-Tools CP-SAT to validate soft constraints.",
    )
    def test_generate_minimizes_teacher_intraday_gaps(self):
        """F-29: solver should prefer compact schedules over fragmented ones."""
        slot_pref_index = build_slot_preference_index(slots=build_weekly_slots())
        allowed = {"MON_09:00", "MON_12:00", "MON_13:00"}
        self.teacher.time_preferences = {
            key: TeacherTimePreferenceState.UNAVAILABLE
            for key in slot_pref_index.values()
            if key not in allowed
        }
        self.subject.weekly_hours = 2
        self.subject.save(update_fields=["weekly_hours"])
        self.teacher.save(update_fields=["time_preferences"])

        response = self.generate_schedule()
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        schedules = Schedule.objects.filter(subject=self.subject)
        assigned_keys = {
            slot_preference_key_from_datetime(slot=s.start_time) for s in schedules
        }
        self.assertEqual(len(assigned_keys), 2)
        self.assertNotIn(
            "MON_09:00",
            assigned_keys,
            "Expected compact assignment (MON_12:00 + MON_13:00) but got a fragmented one.",
        )

    def test_generate_rejects_group_intraday_gaps(self):
        """F-30: a group's timetable cannot contain intra-day gaps."""
        self.subject.weekly_hours = 1
        self.subject.save(update_fields=["weekly_hours"])

        teacher_2 = Teacher.objects.create(
            team=self.team,
            name="Lucia Lopez",
            max_weekly_hours=20,
            working_hours=8,
        )
        Subject.objects.create(
            team=self.team,
            name="Science",
            weekly_hours=1,
            duration=1.0,

            type=SubjectType.NORMAL,
            teacher=teacher_2,
            group=self.group,
        )

        slot_pref_index = build_slot_preference_index(slots=build_weekly_slots())
        allowed = {"MON_08:30", "MON_13:00"}

        self.teacher.time_preferences = {
            key: TeacherTimePreferenceState.UNAVAILABLE
            for key in slot_pref_index.values()
            if key not in allowed
        }
        teacher_2.time_preferences = {
            key: TeacherTimePreferenceState.UNAVAILABLE
            for key in slot_pref_index.values()
            if key not in allowed
        }
        self.teacher.save(update_fields=["time_preferences"])
        teacher_2.save(update_fields=["time_preferences"])

        response = self.generate_schedule()

        self.assert_generate_bad_request_with_detail(
            response,
            "Could not generate a feasible schedule",
        )

    def test_analyze_without_params(self):
        """Test that analyze endpoint requires schedule_ids or source parameter."""
        response = self.client.post(reverse("schedule-analyze"), {}, format="json")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data["_error"]["code"], "INVALID_ANALYZE_PARAMS")

    def test_analyze_with_no_matching_schedules(self):
        """Test that analyze returns error when no schedules match the criteria."""
        response = self.client.post(
            reverse("schedule-analyze"),
            {"schedule_ids": [9999]},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data["_error"]["code"], "NO_SCHEDULES_FOUND")

    def test_analyze_with_specific_schedule_ids(self):
        """Test that analyze works with specific schedule IDs."""
        schedule = self.create_schedule()

        response = self.client.post(
            reverse("schedule-analyze"),
            {"schedule_ids": [schedule.id]},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("count", response.data)
        self.assertIn("defects", response.data)
        self.assertEqual(response.data["count"], len(response.data["defects"]))

    def test_analyze_with_generated_source(self):
        """Test that analyze can filter by generated source."""
        self.create_schedule(
            observations=AUTO_GENERATED_OBSERVATION,
            created_by=self.user.email,
            updated_by=self.user.email,
        )

        response = self.client.post(
            reverse("schedule-analyze"),
            {"source": "generated"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("defects", response.data)

    def test_analyze_with_saved_source(self):
        """Test that analyze can filter by saved source."""
        self.create_schedule(name="Manual Schedule")

        response = self.client.post(
            reverse("schedule-analyze"),
            {"source": "saved"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("defects", response.data)

    def test_analyze_detects_internal_gaps(self):
        """Test that analyze detects internal gaps (gaps between consecutive hours)."""
        today = timezone.now().date()

        # First session: 9:00-10:00
        schedule1 = Schedule.objects.create(
            name="Morning Session",
            start_time=timezone.make_aware(
                datetime.combine(today, datetime.min.time()).replace(hour=9)
            ),
            end_time=timezone.make_aware(
                datetime.combine(today, datetime.min.time()).replace(hour=10)
            ),
            teacher=self.teacher,
            classroom=self.classroom,
            group=self.group,
            subject=self.subject,
            team=self.team,
            created_by=self.user.email,
            updated_by=self.user.email,
        )
        schedule1.users.add(self.user)

        # Second session: 12:00-13:00 (gap at 11:00)
        schedule2 = Schedule.objects.create(
            name="Afternoon Session",
            start_time=timezone.make_aware(
                datetime.combine(today, datetime.min.time()).replace(hour=12)
            ),
            end_time=timezone.make_aware(
                datetime.combine(today, datetime.min.time()).replace(hour=13)
            ),
            teacher=self.teacher,
            classroom=self.classroom,
            group=self.group,
            subject=self.subject,
            team=self.team,
            created_by=self.user.email,
            updated_by=self.user.email,
        )
        schedule2.users.add(self.user)

        response = self.client.post(
            reverse("schedule-analyze"),
            {"schedule_ids": [schedule1.id, schedule2.id]},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertGreater(response.data["count"], 0)

        defects = response.data["defects"]
        internal_gaps = [d for d in defects if d.get("gap_type") == "INTERNAL"]
        self.assertGreater(len(internal_gaps), 0)

        internal_gap = internal_gaps[0]
        self.assertEqual(internal_gap["entity_type"], "group")
        self.assertEqual(internal_gap["entity_id"], self.group.id)
        self.assertEqual(internal_gap["severity"], "MEDIUM")
        self.assertIn("Hueco detectado", internal_gap["description"])

    def test_analyze_detects_boundary_gaps(self):
        """Test that analyze detects boundary gaps (missing sessions in expected range)."""
        today = timezone.now().date()

        # Create one session at 9:00 for a primary group
        # Primary groups should have sessions from 9:00 to 13:00
        # So missing 10:00, 11:30-12:00 (break), 12:00, 13:00 would be boundary gaps
        schedule = Schedule.objects.create(
            name="Single Morning Session",
            start_time=timezone.make_aware(
                datetime.combine(today, datetime.min.time()).replace(hour=9)
            ),
            end_time=timezone.make_aware(
                datetime.combine(today, datetime.min.time()).replace(hour=10)
            ),
            teacher=self.teacher,
            classroom=self.classroom,
            group=self.group,
            subject=self.subject,
            team=self.team,
            created_by=self.user.email,
            updated_by=self.user.email,
        )
        schedule.users.add(self.user)

        response = self.client.post(
            reverse("schedule-analyze"),
            {"schedule_ids": [schedule.id]},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertGreater(response.data["count"], 0)

        defects = response.data["defects"]
        boundary_gaps = [d for d in defects if d.get("gap_type") == "BOUNDARY"]
        self.assertGreater(len(boundary_gaps), 0)

        boundary_gap = boundary_gaps[0]
        self.assertEqual(boundary_gap["entity_type"], "group")
        self.assertEqual(boundary_gap["severity"], "LOW")
        self.assertIn("Sesión faltante", boundary_gap["description"])

    def test_analyze_respects_stage_specific_hours(self):
        """Test that analyze respects different hours for different educational stages."""
        secondary_group = Group.objects.create(
            name="3A ESO",
            stage=EducationalStage.SECONDARY,
            team=self.team,
        )

        secondary_subject = Subject.objects.create(
            team=self.team,
            name="Physics",
            weekly_hours=2,
            duration=1.0,

            type=SubjectType.NORMAL,
            teacher=self.teacher,
            group=secondary_group,
        )

        today = timezone.now().date()

        # ESO starts at 8:00, so create session at 8:00-9:00
        schedule = Schedule.objects.create(
            name="Secondary Session",
            start_time=timezone.make_aware(
                datetime.combine(today, datetime.min.time()).replace(hour=8)
            ),
            end_time=timezone.make_aware(
                datetime.combine(today, datetime.min.time()).replace(hour=9)
            ),
            teacher=self.teacher,
            classroom=self.classroom,
            group=secondary_group,
            subject=secondary_subject,
            team=self.team,
            created_by=self.user.email,
            updated_by=self.user.email,
        )
        schedule.users.add(self.user)

        response = self.client.post(
            reverse("schedule-analyze"),
            {"schedule_ids": [schedule.id]},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        defects = response.data["defects"]
        # Should have boundary gaps for missing hours (9-13)
        # but NOT a gap at 8:00 since ESO starts at 8:00
        descriptions = [d["description"] for d in defects]

        # Should NOT complain about 8:00 being outside expected range
        self.assertFalse(
            any("08:00" in desc and "Sesión faltante" in desc for desc in descriptions),
            "Secondary stage should not report 08:00 as missing",
        )


class ScheduleSlotConfigurationTests(AuthenticatedAdminAPIMixin, APITestCase):
    def setUp(self):
        self.authenticate_admin(email_prefix="schedule-slot-config")
        self.teacher = Teacher.objects.create(
            team=self.team,
            name="Ana Perez",
            max_weekly_hours=40,
            working_hours=12,
        )
        self.classroom = Classroom.objects.create(name="Aula 1A", team=self.team)
        self.group = Group.objects.create(
            name="1A",
            stage=EducationalStage.PRIMARY,
            team=self.team,
        )
        self.subject = Subject.objects.create(
            team=self.team,
            name="Mathematics",
            weekly_hours=25,
            duration=1.0,

            type=SubjectType.NORMAL,
            teacher=self.teacher,
            group=self.group,
        )
        self.subject.mandatory_classroom = self.classroom
        self.subject.save(update_fields=["mandatory_classroom"])

    def test_build_windows_from_stage_config_keeps_primary_half_slot_split_by_break(
        self,
    ):
        windows = build_windows_from_stage_config(
            {
                "start_time": "09:00",
                "end_time": "14:00",
                "breaks": [{"start": "11:30", "end": "12:00"}],
                "session_duration": 60,
            }
        )

        self.assertEqual(
            windows,
            [
                (
                    datetime.strptime("09:00", "%H:%M").time(),
                    datetime.strptime("10:00", "%H:%M").time(),
                    False,
                ),
                (
                    datetime.strptime("10:00", "%H:%M").time(),
                    datetime.strptime("11:00", "%H:%M").time(),
                    False,
                ),
                (
                    datetime.strptime("11:00", "%H:%M").time(),
                    datetime.strptime("11:30", "%H:%M").time(),
                    False,
                ),
                (
                    datetime.strptime("11:30", "%H:%M").time(),
                    datetime.strptime("12:00", "%H:%M").time(),
                    True,
                ),
                (
                    datetime.strptime("12:00", "%H:%M").time(),
                    datetime.strptime("13:00", "%H:%M").time(),
                    False,
                ),
                (
                    datetime.strptime("13:00", "%H:%M").time(),
                    datetime.strptime("14:00", "%H:%M").time(),
                    False,
                ),
            ],
        )

    def test_build_windows_from_stage_config_keeps_preschool_half_slots_split_by_breaks(
        self,
    ):
        windows = build_windows_from_stage_config(
            {
                "start_time": "09:00",
                "end_time": "14:00",
                "breaks": [
                    {"start": "10:30", "end": "11:00"},
                    {"start": "13:30", "end": "14:00"},
                ],
                "session_duration": 60,
            }
        )

        self.assertEqual(
            windows,
            [
                (
                    datetime.strptime("09:00", "%H:%M").time(),
                    datetime.strptime("10:00", "%H:%M").time(),
                    False,
                ),
                (
                    datetime.strptime("10:00", "%H:%M").time(),
                    datetime.strptime("10:30", "%H:%M").time(),
                    False,
                ),
                (
                    datetime.strptime("10:30", "%H:%M").time(),
                    datetime.strptime("11:00", "%H:%M").time(),
                    True,
                ),
                (
                    datetime.strptime("11:00", "%H:%M").time(),
                    datetime.strptime("12:00", "%H:%M").time(),
                    False,
                ),
                (
                    datetime.strptime("12:00", "%H:%M").time(),
                    datetime.strptime("13:00", "%H:%M").time(),
                    False,
                ),
                (
                    datetime.strptime("13:00", "%H:%M").time(),
                    datetime.strptime("13:30", "%H:%M").time(),
                    False,
                ),
                (
                    datetime.strptime("13:30", "%H:%M").time(),
                    datetime.strptime("14:00", "%H:%M").time(),
                    True,
                ),
            ],
        )

    def test_stage_window_diagnostics_do_not_flag_primary_config_split_by_recess(self):
        self.team.schedule_config = {
            "PRIMARY": {
                "label": "Primaria",
                "color": "blue",
                "start_time": "09:00",
                "end_time": "14:00",
                "breaks": [{"start": "11:30", "end": "12:00"}],
                "session_duration": 60,
            }
        }
        self.team.save(update_fields=["schedule_config"])
        slots = build_weekly_slots(
            stage_slot_windows=parse_schedule_config_to_slot_windows(
                self.team.schedule_config
            )
        )
        sessions = BasicScheduleGenerator._build_sessions(
            subjects=[self.subject],
            fallback_teacher=self.teacher,
        )

        diagnostics = collect_generation_diagnostics(
            subjects=[self.subject],
            teachers=[self.teacher],
            sessions=sessions,
            slots=slots,
            classrooms=[self.classroom],
            generation_options={},
        )

        self.assertNotIn(
            "STAGE_SLOT_WINDOW_TOO_NARROW",
            [entry["code"] for entry in diagnostics],
        )

    @skipIf(
        schedule_assignment.cp_model is None,
        "Requires OR-Tools CP-SAT to validate slot overlap constraints.",
    )
    def test_solver_prevents_overlap_between_30_and_60_minute_slots(self):
        other_group = Group.objects.create(
            name="2A",
            stage=EducationalStage.PRIMARY,
            team=self.team,
        )
        other_subject = Subject.objects.create(
            team=self.team,
            name="Science",
            weekly_hours=1,
            duration=1.0,

            type=SubjectType.NORMAL,
            teacher=self.teacher,
            group=other_group,
        )
        other_subject.mandatory_classroom = self.classroom
        other_subject.save(update_fields=["mandatory_classroom"])

        sessions = [
            {
                "teacher": self.teacher,
                "teacher_id": self.teacher.id,
                "group": self.group,
                "subject": self.subject,
                "allowed_classroom_ids": {self.classroom.id},
                "name": self.subject.name,
            },
            {
                "teacher": self.teacher,
                "teacher_id": self.teacher.id,
                "group": other_group,
                "subject": other_subject,
                "allowed_classroom_ids": {self.classroom.id},
                "name": other_subject.name,
            },
        ]
        today = timezone.localdate()
        slots = [
            {
                "start": timezone.make_aware(
                    datetime.combine(today, datetime.min.time()).replace(hour=11)
                ),
                "end": timezone.make_aware(
                    datetime.combine(today, datetime.min.time()).replace(hour=12)
                ),
                "stage": STAGE_PRIMARY,
                "day_code": "MON",
                "is_recess": False,
            },
            {
                "start": timezone.make_aware(
                    datetime.combine(today, datetime.min.time()).replace(hour=11)
                ),
                "end": timezone.make_aware(
                    datetime.combine(today, datetime.min.time()).replace(
                        hour=11, minute=30
                    )
                ),
                "stage": STAGE_PRIMARY,
                "day_code": "MON",
                "is_recess": False,
            },
            {
                "start": timezone.make_aware(
                    datetime.combine(today, datetime.min.time()).replace(hour=12)
                ),
                "end": timezone.make_aware(
                    datetime.combine(today, datetime.min.time()).replace(hour=13)
                ),
                "stage": STAGE_PRIMARY,
                "day_code": "MON",
                "is_recess": False,
            },
        ]

        compatible_classrooms_by_session = (
            schedule_assignment._build_compatible_classroom_index(
                sessions=sessions,
                classrooms=[self.classroom],
            )
        )
        slot_by_session, _classroom_by_session, _, _ = (
            schedule_assignment._cp_sat_session_assignment(
                sessions=sessions,
                slots=slots,
                compatible_classrooms_by_session=compatible_classrooms_by_session,
                random_seed=None,
                fixed_assignments=None,
                previous_assignment_by_session=None,
                generation_options=None,
            )
        )

        assigned_ranges = [
            (slots[slot_index]["start"], slots[slot_index]["end"])
            for slot_index in slot_by_session
        ]
        self.assertEqual(len(assigned_ranges), 2)
        left_start, left_end = assigned_ranges[0]
        right_start, right_end = assigned_ranges[1]
        self.assertFalse(left_start < right_end and right_start < left_end)


# ---------------------------------------------------------------------------
# TC Tests — helpers
# ---------------------------------------------------------------------------


def _make_team_and_teacher(
    *, email_prefix, max_weekly_hours=20, weekly_hours_exact=False
):
    """Create an isolated team + teacher for unit tests that don't use APITestCase."""
    from user.models import CollaborationTeam

    team = CollaborationTeam.objects.create(name=f"{email_prefix}-team")
    teacher = Teacher.objects.create(
        team=team,
        name=f"Teacher {email_prefix}",
        max_weekly_hours=max_weekly_hours,
        weekly_hours_exact=weekly_hours_exact,
    )
    return team, teacher


def _next_monday_slot(hour_start=9, hour_end=10):
    """Return (start_dt, end_dt) for next Monday at the given hours (timezone-aware)."""
    now = timezone.localtime()
    days_until = (7 - now.weekday()) % 7 or 7
    monday = now.date() + timedelta(days=days_until)
    start_dt = timezone.make_aware(
        datetime.combine(monday, datetime.min.time().replace(hour=hour_start))
    )
    end_dt = timezone.make_aware(
        datetime.combine(monday, datetime.min.time().replace(hour=hour_end))
    )
    return start_dt, end_dt


# ---------------------------------------------------------------------------
# TC Model tests
# ---------------------------------------------------------------------------


class TestTCSessionModel(TestCase):
    def setUp(self):
        self.team, self.teacher = _make_team_and_teacher(email_prefix="tc-model")

    def test_creacion_basica(self):
        from datetime import time

        tc = TCSession.objects.create(
            teacher=self.teacher,
            team=self.team,
            day=0,
            start_time=time(9, 0),
            end_time=time(10, 0),
        )
        self.assertEqual(tc.day, 0)
        self.assertEqual(tc.teacher, self.teacher)
        self.assertEqual(tc.team, self.team)

    def test_str(self):
        from datetime import time

        tc = TCSession(
            teacher=self.teacher,
            team=self.team,
            day=0,
            start_time=time(9, 0),
            end_time=time(10, 0),
        )
        self.assertIn("Lunes", str(tc))
        self.assertIn(self.teacher.name, str(tc))

    def test_ordering(self):
        from datetime import time

        TCSession.objects.create(
            teacher=self.teacher,
            team=self.team,
            day=1,
            start_time=time(10, 0),
            end_time=time(11, 0),
        )
        TCSession.objects.create(
            teacher=self.teacher,
            team=self.team,
            day=0,
            start_time=time(9, 0),
            end_time=time(10, 0),
        )
        sessions = list(TCSession.objects.filter(team=self.team))
        self.assertEqual(sessions[0].day, 0)
        self.assertEqual(sessions[1].day, 1)


# ---------------------------------------------------------------------------
# TC Assigner unit tests
# ---------------------------------------------------------------------------


class TestTCAssigner(TestCase):
    def setUp(self):
        self.team, self.teacher = _make_team_and_teacher(email_prefix="tc-algo")
        self.slots = build_weekly_slots()

    def _make_unsaved_schedule(self, slot_index):
        """Return a lightweight object mimicking an unsaved Schedule for tc_assigner input."""
        slot = self.slots[slot_index]
        return SimpleNamespace(
            teacher=self.teacher,
            teacher_id=self.teacher.id,
            start_time=slot["start"],
            end_time=slot["end"],
        )

    def _first_non_recess_slot_index(self):
        return next(i for i, s in enumerate(self.slots) if not s.get("is_recess"))

    def test_teachers_on_duty_cero_no_genera_nada(self):
        result = assign_tc_sessions(
            teachers=[self.teacher],
            existing_schedules=[],
            weekly_slots=self.slots,
            teachers_on_duty=0,
            team=self.team,
        )
        self.assertEqual(result.tc_sessions, [])
        self.assertEqual(result.warnings, [])

    def test_docente_con_clase_excluido(self):
        idx = self._first_non_recess_slot_index()
        slot = self.slots[idx]
        unsaved = self._make_unsaved_schedule(idx)

        result = assign_tc_sessions(
            teachers=[self.teacher],
            existing_schedules=[unsaved],
            weekly_slots=self.slots,
            teachers_on_duty=1,
            team=self.team,
        )
        # Teacher has class at this slot — should not get a TCSession there
        slot_day = slot["start"].weekday()
        slot_time = slot["start"].time()
        conflict = any(
            tc.day == slot_day and tc.start_time == slot_time
            for tc in result.tc_sessions
        )
        self.assertFalse(conflict)

    def test_docente_con_clase_solapada_excluido(self):
        """TC at 11:30 must be blocked when teacher has class 11:00-12:00 (partial overlap)."""
        from datetime import datetime, time, timezone

        ref = datetime(2024, 1, 1, tzinfo=timezone.utc)  # Monday
        schedule = SimpleNamespace(
            teacher=self.teacher,
            teacher_id=self.teacher.id,
            start_time=ref.replace(hour=11, minute=0),
            end_time=ref.replace(hour=12, minute=0),
        )
        # Build a TC slot at 11:30–12:00 manually
        tc_slot = SimpleNamespace(
            day=0,
            start_time=time(11, 30),
            end_time=time(12, 0),
        )
        from schedule.algorithm.tc_assigner import (
            _compute_busy_intervals,
            _overlaps_any,
        )

        busy = _compute_busy_intervals([schedule])
        self.assertTrue(
            _overlaps_any(
                tc_slot.day,
                tc_slot.start_time,
                tc_slot.end_time,
                busy.get(self.teacher.id, []),
            ),
            "Partial overlap not detected: class 11:00-12:00 should block TC at 11:30-12:00",
        )

    def test_docente_unavailable_excluido(self):
        idx = self._first_non_recess_slot_index()
        slot = self.slots[idx]
        key = slot_instance_key(slot=slot)
        self.teacher.time_preferences = {key: TeacherTimePreferenceState.UNAVAILABLE}
        self.teacher.save()

        result = assign_tc_sessions(
            teachers=[self.teacher],
            existing_schedules=[],
            weekly_slots=self.slots,
            teachers_on_duty=1,
            team=self.team,
        )
        slot_day = slot["start"].weekday()
        slot_time = slot["start"].time()
        conflict = any(
            tc.day == slot_day and tc.start_time == slot_time
            for tc in result.tc_sessions
        )
        self.assertFalse(conflict)

    def test_warning_cuando_hay_pocos_candidatos(self):
        result = assign_tc_sessions(
            teachers=[self.teacher],
            existing_schedules=[],
            weekly_slots=self.slots,
            teachers_on_duty=2,  # only 1 teacher available
            team=self.team,
        )
        self.assertGreater(len(result.warnings), 0)
        first_warning = result.warnings[0]
        self.assertEqual(first_warning["required"], 2)
        self.assertEqual(first_warning["assigned"], 1)

    def test_sin_warning_cuando_cobertura_completa(self):
        self.teacher.max_weekly_hours = 100
        self.teacher.save()
        _, teacher2 = _make_team_and_teacher(
            email_prefix="tc-algo2", max_weekly_hours=100
        )
        teacher2.team = self.team
        teacher2.save()

        result = assign_tc_sessions(
            teachers=[self.teacher, teacher2],
            existing_schedules=[],
            weekly_slots=self.slots,
            teachers_on_duty=2,
            team=self.team,
        )
        self.assertEqual(result.warnings, [])

    def test_deduplicacion_slots_multietapa(self):
        result = assign_tc_sessions(
            teachers=[self.teacher],
            existing_schedules=[],
            weekly_slots=self.slots,
            teachers_on_duty=1,
            team=self.team,
        )
        # No TCSession should share (day, start_time) — deduplicated across stages
        seen = set()
        for tc in result.tc_sessions:
            key = (tc.day, tc.start_time)
            self.assertNotIn(
                key,
                seen,
                "Overlapping TCSession found: same (day, start_time) assigned twice",
            )
            seen.add(key)

    def test_invariante_no_solapamiento_schedule_tcsession(self):
        idx = self._first_non_recess_slot_index()
        unsaved = self._make_unsaved_schedule(idx)

        result = assign_tc_sessions(
            teachers=[self.teacher],
            existing_schedules=[unsaved],
            weekly_slots=self.slots,
            teachers_on_duty=1,
            team=self.team,
        )
        busy_slots = {(s.start_time.weekday(), s.start_time.time()) for s in [unsaved]}
        for tc in result.tc_sessions:
            self.assertNotIn(
                (tc.day, tc.start_time),
                busy_slots,
                "TCSession overlaps with a Schedule slot",
            )

    def test_hueco_muerto_priorizado(self):
        """Teacher with a dead gap (class before AND after) should appear before one without."""
        _, teacher_no_gap = _make_team_and_teacher(email_prefix="tc-no-gap")
        teacher_no_gap.team = self.team
        teacher_no_gap.save()

        # Give teacher_with_gap a class at slot 0 and slot 2 of Monday → slot 1 is a dead gap
        non_recess = [s for s in self.slots if not s.get("is_recess")]
        monday_slots = [s for s in non_recess if s["start"].weekday() == 0]
        if len(monday_slots) < 3:
            self.skipTest("Not enough Monday slots to test dead gap")

        def _unsaved(slot):
            return SimpleNamespace(
                teacher=self.teacher,
                teacher_id=self.teacher.id,
                start_time=slot["start"],
                end_time=slot["end"],
            )

        existing = [_unsaved(monday_slots[0]), _unsaved(monday_slots[2])]
        gap_slot = monday_slots[1]

        result = assign_tc_sessions(
            teachers=[self.teacher, teacher_no_gap],
            existing_schedules=existing,
            weekly_slots=self.slots,
            teachers_on_duty=1,
            team=self.team,
        )
        gap_day = gap_slot["start"].weekday()
        gap_time = gap_slot["start"].time()
        gap_tc = [
            tc
            for tc in result.tc_sessions
            if tc.day == gap_day and tc.start_time == gap_time
        ]
        if gap_tc:
            self.assertEqual(gap_tc[0].teacher_id, self.teacher.id)


# ---------------------------------------------------------------------------
# TC API view tests
# ---------------------------------------------------------------------------


class TestTCSessionListView(AuthenticatedAdminAPIMixin, APITestCase):
    def setUp(self):
        self.authenticate_admin(email_prefix="tc-list")
        self.teacher = Teacher.objects.create(
            team=self.team, name="Profe Lista", max_weekly_hours=18
        )

    def _create_tc(self, day=0):
        from datetime import time

        return TCSession.objects.create(
            teacher=self.teacher,
            team=self.team,
            day=day,
            start_time=time(9, 0),
            end_time=time(10, 0),
        )

    def test_lista_solo_sesiones_del_equipo(self):
        self._create_tc()
        _, other_team = self.create_isolated_user(email_prefix="tc-other")
        other_teacher = Teacher.objects.create(
            team=other_team, name="Otro", max_weekly_hours=10
        )
        from datetime import time

        TCSession.objects.create(
            teacher=other_teacher,
            team=other_team,
            day=0,
            start_time=time(9, 0),
            end_time=time(10, 0),
        )
        response = self.client.get(reverse("tc-session-list"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)

    def test_filtro_por_teacher(self):
        self._create_tc()
        teacher2 = Teacher.objects.create(
            team=self.team, name="Profe2", max_weekly_hours=18
        )
        from datetime import time

        TCSession.objects.create(
            teacher=teacher2,
            team=self.team,
            day=1,
            start_time=time(10, 0),
            end_time=time(11, 0),
        )
        response = self.client.get(
            reverse("tc-session-list"), {"teacher": self.teacher.id}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["teacher"], self.teacher.id)

    def test_filtro_por_day(self):
        self._create_tc(day=0)
        self._create_tc(day=2)
        response = self.client.get(reverse("tc-session-list"), {"day": 2})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["day"], 2)

    def test_requiere_autenticacion(self):
        self.client.force_authenticate(None)
        response = self.client.get(reverse("tc-session-list"))
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)


class TestTCSessionCreateView(AuthenticatedAdminAPIMixin, APITestCase):
    def setUp(self):
        self.authenticate_admin(email_prefix="tc-create")
        self.teacher = Teacher.objects.create(
            team=self.team, name="Profe Create", max_weekly_hours=18
        )
        self.classroom = Classroom.objects.create(name="Aula", team=self.team)
        self.group = Group.objects.create(
            name="1A", stage=EducationalStage.PRIMARY, team=self.team
        )

    def _payload(self, **overrides):
        data = {
            "teacher": self.teacher.id,
            "day": 0,
            "start_time": "09:00",
            "end_time": "10:00",
        }
        data.update(overrides)
        return data

    def test_crear_tc_session_ok(self):
        response = self.client.post(
            reverse("tc-session-create"), self._payload(), format="json"
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertIn("tc_session", response.data)
        self.assertEqual(TCSession.objects.filter(team=self.team).count(), 1)

    def test_crear_tc_session_conflicto_schedule(self):
        start_dt, end_dt = _next_monday_slot(9, 10)
        sch = Schedule.objects.create(
            name="Math",
            start_time=start_dt,
            end_time=end_dt,
            teacher=self.teacher,
            classroom=self.classroom,
            group=self.group,
            team=self.team,
            created_by="test",
            updated_by="test",
            observations="",
        )
        sch.users.add(self.user)

        response = self.client.post(
            reverse("tc-session-create"),
            self._payload(day=0, start_time="09:00", end_time="10:00"),
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_409_CONFLICT)

    def test_crear_tc_session_conflicto_tc_existente(self):
        from datetime import time

        TCSession.objects.create(
            teacher=self.teacher,
            team=self.team,
            day=0,
            start_time=time(9, 0),
            end_time=time(10, 0),
        )
        response = self.client.post(
            reverse("tc-session-create"), self._payload(), format="json"
        )
        self.assertEqual(response.status_code, status.HTTP_409_CONFLICT)

    def test_crear_tc_session_unavailable(self):
        slots = build_weekly_slots()
        monday_slot = next(
            s
            for s in slots
            if s["start"].weekday() == 0
            and s["start"].time().hour == 9
            and not s.get("is_recess")
        )
        key = slot_instance_key(slot=monday_slot)
        self.teacher.time_preferences = {key: TeacherTimePreferenceState.UNAVAILABLE}
        self.teacher.save()

        response = self.client.post(
            reverse("tc-session-create"), self._payload(), format="json"
        )
        self.assertEqual(response.status_code, status.HTTP_409_CONFLICT)

    def test_crear_tc_session_warning_horas_exactas(self):
        teacher_exact = Teacher.objects.create(
            team=self.team,
            name="Profe Exact",
            max_weekly_hours=1,
            weekly_hours_exact=True,
        )
        # Create 1h TC to fill the quota, then add another
        from datetime import time

        TCSession.objects.create(
            teacher=teacher_exact,
            team=self.team,
            day=1,
            start_time=time(9, 0),
            end_time=time(10, 0),
        )
        response = self.client.post(
            reverse("tc-session-create"),
            {
                "teacher": teacher_exact.id,
                "day": 0,
                "start_time": "09:00",
                "end_time": "10:00",
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertIn("warning", response.data)


class TestTCSessionDeleteView(AuthenticatedAdminAPIMixin, APITestCase):
    def setUp(self):
        self.authenticate_admin(email_prefix="tc-delete")
        self.teacher = Teacher.objects.create(
            team=self.team, name="Profe Delete", max_weekly_hours=18
        )

    def _create_tc(self, **kwargs):
        from datetime import time

        defaults = dict(
            teacher=self.teacher,
            team=self.team,
            day=0,
            start_time=time(9, 0),
            end_time=time(10, 0),
        )
        defaults.update(kwargs)
        return TCSession.objects.create(**defaults)

    def test_eliminar_tc_session_ok(self):
        tc = self._create_tc()
        response = self.client.delete(
            reverse("tc-session-delete", kwargs={"pk": tc.pk})
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data["deleted"])
        self.assertFalse(TCSession.objects.filter(pk=tc.pk).exists())

    def test_eliminar_tc_session_warning_horas_exactas(self):
        teacher_exact = Teacher.objects.create(
            team=self.team,
            name="Exact",
            max_weekly_hours=2,
            weekly_hours_exact=True,
        )
        from datetime import time

        tc = TCSession.objects.create(
            teacher=teacher_exact,
            team=self.team,
            day=0,
            start_time=time(9, 0),
            end_time=time(10, 0),
        )
        response = self.client.delete(
            reverse("tc-session-delete", kwargs={"pk": tc.pk})
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("warning", response.data)

    def test_eliminar_tc_session_otro_equipo(self):
        _, other_team = self.create_isolated_user(email_prefix="tc-del-other")
        other_teacher = Teacher.objects.create(
            team=other_team, name="Otro", max_weekly_hours=10
        )
        from datetime import time

        tc = TCSession.objects.create(
            teacher=other_teacher,
            team=other_team,
            day=0,
            start_time=time(9, 0),
            end_time=time(10, 0),
        )
        response = self.client.delete(
            reverse("tc-session-delete", kwargs={"pk": tc.pk})
        )
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)


class TestTCSessionSwapView(AuthenticatedAdminAPIMixin, APITestCase):
    def setUp(self):
        self.authenticate_admin(email_prefix="tc-swap")
        self.teacher_a = Teacher.objects.create(
            team=self.team, name="Profe A", max_weekly_hours=18
        )
        self.teacher_b = Teacher.objects.create(
            team=self.team, name="Profe B", max_weekly_hours=18
        )
        self.classroom = Classroom.objects.create(name="Aula", team=self.team)
        self.group = Group.objects.create(
            name="1A", stage=EducationalStage.PRIMARY, team=self.team
        )

    def _create_tc(self, teacher, day, hour_start=9, hour_end=10):
        from datetime import time

        return TCSession.objects.create(
            teacher=teacher,
            team=self.team,
            day=day,
            start_time=time(hour_start, 0),
            end_time=time(hour_end, 0),
        )

    def _swap(self, tc_a, tc_b):
        return self.client.post(
            reverse("tc-session-swap"),
            {"tc_session_a": tc_a.pk, "tc_session_b": tc_b.pk},
            format="json",
        )

    def test_swap_ok(self):
        tc_a = self._create_tc(self.teacher_a, day=0)
        tc_b = self._create_tc(self.teacher_b, day=1)
        response = self._swap(tc_a, tc_b)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        tc_a.refresh_from_db()
        tc_b.refresh_from_db()
        self.assertEqual(tc_a.day, 1)
        self.assertEqual(tc_b.day, 0)

    def test_swap_mismo_docente_distinto_slot(self):
        tc_a = self._create_tc(self.teacher_a, day=0)
        tc_b = self._create_tc(self.teacher_a, day=1)
        response = self._swap(tc_a, tc_b)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        tc_a.refresh_from_db()
        self.assertEqual(tc_a.day, 1)

    def test_swap_distinta_duracion(self):
        tc_a = self._create_tc(self.teacher_a, day=0, hour_start=9, hour_end=10)
        tc_b = self._create_tc(self.teacher_b, day=1, hour_start=9, hour_end=11)
        response = self._swap(tc_a, tc_b)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_swap_conflicto_schedule_docente_a(self):
        tc_a = self._create_tc(self.teacher_a, day=0)
        tc_b = self._create_tc(self.teacher_b, day=1)
        # Give teacher_a a class at slot B (day=1, 09:00)
        start_dt, end_dt = _next_monday_slot(9, 10)
        tuesday = start_dt + timedelta(days=1)
        tuesday_end = end_dt + timedelta(days=1)
        sch = Schedule.objects.create(
            name="Conflict",
            start_time=tuesday,
            end_time=tuesday_end,
            teacher=self.teacher_a,
            classroom=self.classroom,
            group=self.group,
            team=self.team,
            created_by="test",
            updated_by="test",
            observations="",
        )
        sch.users.add(self.user)
        response = self._swap(tc_a, tc_b)
        self.assertEqual(response.status_code, status.HTTP_409_CONFLICT)

    def test_swap_conflicto_schedule_docente_b(self):
        tc_a = self._create_tc(self.teacher_a, day=0)
        tc_b = self._create_tc(self.teacher_b, day=1)
        # Give teacher_b a class at slot A (day=0, 09:00)
        start_dt, end_dt = _next_monday_slot(9, 10)
        sch = Schedule.objects.create(
            name="Conflict B",
            start_time=start_dt,
            end_time=end_dt,
            teacher=self.teacher_b,
            classroom=self.classroom,
            group=self.group,
            team=self.team,
            created_by="test",
            updated_by="test",
            observations="",
        )
        sch.users.add(self.user)
        response = self._swap(tc_a, tc_b)
        self.assertEqual(response.status_code, status.HTTP_409_CONFLICT)

    def test_swap_unavailable_docente_a(self):
        slots = build_weekly_slots()
        tuesday_slot = next(
            s
            for s in slots
            if s["start"].weekday() == 1
            and s["start"].time().hour == 9
            and not s.get("is_recess")
        )
        key = slot_instance_key(slot=tuesday_slot)
        self.teacher_a.time_preferences = {key: TeacherTimePreferenceState.UNAVAILABLE}
        self.teacher_a.save()

        tc_a = self._create_tc(self.teacher_a, day=0)
        tc_b = self._create_tc(self.teacher_b, day=1)
        response = self._swap(tc_a, tc_b)
        self.assertEqual(response.status_code, status.HTTP_409_CONFLICT)

    def test_swap_unavailable_docente_b(self):
        slots = build_weekly_slots()
        monday_slot = next(
            s
            for s in slots
            if s["start"].weekday() == 0
            and s["start"].time().hour == 9
            and not s.get("is_recess")
        )
        key = slot_instance_key(slot=monday_slot)
        self.teacher_b.time_preferences = {key: TeacherTimePreferenceState.UNAVAILABLE}
        self.teacher_b.save()

        tc_a = self._create_tc(self.teacher_a, day=0)
        tc_b = self._create_tc(self.teacher_b, day=1)
        response = self._swap(tc_a, tc_b)
        self.assertEqual(response.status_code, status.HTTP_409_CONFLICT)

    def test_swap_tc_session_otro_equipo(self):
        _, other_team = self.create_isolated_user(email_prefix="tc-swap-other")
        other_teacher = Teacher.objects.create(
            team=other_team, name="Otro", max_weekly_hours=10
        )
        from datetime import time

        tc_other = TCSession.objects.create(
            teacher=other_teacher,
            team=other_team,
            day=1,
            start_time=time(9, 0),
            end_time=time(10, 0),
        )
        tc_own = self._create_tc(self.teacher_a, day=0)
        response = self._swap(tc_own, tc_other)
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)
